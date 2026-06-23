# -*- coding: utf-8 -*-
"""关键词库 service（06-23-tab）.

CRUD + excel 导入（合并追加去重）+ excel 导出。

* excel 导入规范：表头 ``卡片名称,关键词,风险等级,备注``。一行一个关键词。卡片名称
  相同的行合并为同一张卡（风险等级/备注取该卡片名分组首行非空值）。
* 合并追加去重：同名卡片存在则把新 term 追加进旧卡（已有 term 跳过），
  risk_level/note 用 excel 新值覆盖；同名卡片不存在则新建。
* 导入返统计：新建卡片数 / 追加卡片数 / 新增词数 / 跳过重复词数 / 拒绝行数
  （风险等级非法值的行跳过并记入 rejected）。
* 导出：返 xlsx 流，表头 ``卡片名称,关键词,风险等级,备注``，一行一词，卡片名连续多行。
"""

import io
import logging
from typing import Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import KeywordCard, KeywordHit, KeywordTerm
from app.models.keyword import RISK_LEVELS, RISK_MEDIUM

logger = logging.getLogger(__name__)


# excel 表头（导入/导出一致）。
EXCEL_HEADER_CARD = "卡片名称"
EXCEL_HEADER_TERM = "关键词"
EXCEL_HEADER_RISK = "风险等级"
EXCEL_HEADER_NOTE = "备注"
EXCEL_HEADERS = [EXCEL_HEADER_CARD, EXCEL_HEADER_TERM, EXCEL_HEADER_RISK, EXCEL_HEADER_NOTE]

# _parse_row 哨兵：非法风险等级行返此对象，区别于 None（整行空/无 term 的静默跳过）。
_REJECTED_ROW = object()


class KeywordLibraryService:
    """Keyword library CRUD + excel import/export."""

    async def list_cards(self, db: AsyncSession) -> list[dict]:
        """列出所有卡片（含每卡 term 数 + 风险等级 + 备注）。

        返回 ``[{id, name, risk_level, note, term_count, created_at, updated_at}]``。
        """
        # 子查询：每卡 term 数。
        count_subq = (
            select(KeywordTerm.card_id, func.count(KeywordTerm.id).label("cnt"))
            .group_by(KeywordTerm.card_id)
            .subquery()
        )
        result = await db.execute(
            select(KeywordCard, count_subq.c.cnt)
            .outerjoin(count_subq, KeywordCard.id == count_subq.c.card_id)
            .order_by(KeywordCard.id.asc())
        )
        rows: list[dict] = []
        for card, cnt in result.all():
            rows.append(
                {
                    "id": card.id,
                    "name": card.name,
                    "risk_level": card.risk_level,
                    "note": card.note,
                    "term_count": int(cnt or 0),
                    "created_at": card.created_at,
                    "updated_at": card.updated_at,
                }
            )
        return rows

    async def get_card(self, db: AsyncSession, card_id: int) -> Optional[dict]:
        """卡片详情（含 terms 列表）。返 ``{id, name, risk_level, note, terms, ...}``。"""
        result = await db.execute(
            select(KeywordCard).where(KeywordCard.id == card_id)
        )
        card = result.scalar_one_or_none()
        if card is None:
            return None
        terms_result = await db.execute(
            select(KeywordTerm)
            .where(KeywordTerm.card_id == card_id)
            .order_by(KeywordTerm.id.asc())
        )
        terms = [
            {"id": t.id, "term": t.term, "created_at": t.created_at}
            for t in terms_result.scalars().all()
        ]
        return {
            "id": card.id,
            "name": card.name,
            "risk_level": card.risk_level,
            "note": card.note,
            "terms": terms,
            "created_at": card.created_at,
            "updated_at": card.updated_at,
        }

    async def create_card(
        self,
        db: AsyncSession,
        name: str,
        risk_level: str,
        note: Optional[str],
        terms: list[str],
    ) -> KeywordCard:
        """新建卡片 + terms（去重保序）。"""
        clean_name = (name or "").strip()
        if not clean_name:
            raise ValueError("卡片名称不能为空")
        if risk_level not in RISK_LEVELS:
            raise ValueError("风险等级必须为 高/中/低")
        card = KeywordCard(
            name=clean_name,
            risk_level=risk_level,
            note=(note or None),
        )
        db.add(card)
        await db.flush()
        for term in self._dedup_terms(terms):
            db.add(KeywordTerm(card_id=card.id, term=term))
        await db.flush()
        logger.info("新建关键词卡片: id=%s name=%s terms=%d", card.id, clean_name, len(terms))
        return card

    async def update_card(
        self,
        db: AsyncSession,
        card_id: int,
        name: Optional[str],
        risk_level: Optional[str],
        note: Optional[str],
        terms: Optional[list[str]],
    ) -> KeywordCard:
        """编辑卡片（name/risk_level/note 可改；terms 全量替换）。"""
        result = await db.execute(select(KeywordCard).where(KeywordCard.id == card_id))
        card = result.scalar_one_or_none()
        if card is None:
            raise LookupError("Keyword card not found")

        if name is not None:
            clean_name = name.strip()
            if not clean_name:
                raise ValueError("卡片名称不能为空")
            card.name = clean_name
        if risk_level is not None:
            if risk_level not in RISK_LEVELS:
                raise ValueError("风险等级必须为 高/中/低")
            card.risk_level = risk_level
        # note 显式传 None → 清空；不传（未提供）→ 不改。
        if note is not None:
            card.note = note.strip() or None

        if terms is not None:
            # 全量替换：删旧 terms 再插新（命中表的 keyword_term_id 会因 CASCADE 失效，
            # 符合「terms 全量替换」语义——历史命中需重跑重建）。
            await db.execute(delete(KeywordTerm).where(KeywordTerm.card_id == card_id))
            for term in self._dedup_terms(terms):
                db.add(KeywordTerm(card_id=card_id, term=term))

        await db.flush()
        logger.info("编辑关键词卡片: id=%s", card_id)
        return card

    async def delete_card(self, db: AsyncSession, card_id: int) -> None:
        """删卡（级联删 terms）。若该卡已被命中引用 → 抛 LookupError 由 router 转 409。

        注：keyword_card_id FK ondelete=RESTRICT，命中引用存在时 DB 层会报约束错；
        这里在应用层提前检查并返明确提示，避免暴露底层 IntegrityError。
        """
        result = await db.execute(select(KeywordCard).where(KeywordCard.id == card_id))
        card = result.scalar_one_or_none()
        if card is None:
            raise LookupError("Keyword card not found")

        hits_result = await db.execute(
            select(KeywordHit).where(KeywordHit.keyword_card_id == card_id).limit(1)
        )
        if hits_result.scalar_one_or_none() is not None:
            raise ValueError("该卡片已被任务审查引用，请先解除关联再删除")

        await db.delete(card)
        logger.info("删除关键词卡片: id=%s name=%s", card_id, card.name)

    # ------------------------------------------------------------------
    # excel 导入 / 导出
    # ------------------------------------------------------------------

    async def import_excel(self, db: AsyncSession, file_bytes: bytes) -> dict:
        """excel 导入（合并追加去重）。

        返回统计 ``{created_cards, appended_cards, new_terms, skipped_terms, rejected_rows}``。
        """
        # openpyxl read_only + data_only，try/finally close（spec 范式）。
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        parsed_rows: list[dict] = []
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                raise ValueError("excel 文件为空或表头缺失")
            col_index = self._parse_header(header)
            rejected = 0
            for row in rows_iter:
                parsed = self._parse_row(row, col_index)
                if parsed is None or parsed is _REJECTED_ROW:
                    if parsed is _REJECTED_ROW:
                        rejected += 1
                    continue
                parsed_rows.append(parsed)
        finally:
            wb.close()

        stats = await self._merge_rows(db, parsed_rows, rejected)
        logger.info(
            "关键词库导入完成: 新建卡片=%d 追加卡片=%d 新增词=%d 跳过=%d 拒绝=%d",
            stats["created_cards"],
            stats["appended_cards"],
            stats["new_terms"],
            stats["skipped_terms"],
            stats["rejected_rows"],
        )
        return stats

    async def export_excel(self, db: AsyncSession) -> bytes:
        """excel 导出：表头 ``卡片名称,关键词,风险等级,备注``，一行一词，卡片名连续多行。"""
        cards = await self.list_cards(db)
        # 取所有 term（按 card_id 分组）。
        card_ids = [c["id"] for c in cards]
        terms_by_card: dict[int, list[str]] = {cid: [] for cid in card_ids}
        if card_ids:
            result = await db.execute(
                select(KeywordTerm)
                .where(KeywordTerm.card_id.in_(card_ids))
                .order_by(KeywordTerm.card_id.asc(), KeywordTerm.id.asc())
            )
            for t in result.scalars().all():
                terms_by_card[t.card_id].append(t.term)

        # openpyxl write — try/finally close（spec 范式）。
        wb = Workbook()
        try:
            ws = wb.active
            ws.append(EXCEL_HEADERS)
            for card in cards:
                terms = terms_by_card.get(card["id"], [])
                if not terms:
                    # 无词的卡片也导出一行（空关键词）以便导入能还原。
                    ws.append([card["name"], "", card["risk_level"], card.get("note") or ""])
                    continue
                for term in terms:
                    ws.append([card["name"], term, card["risk_level"], card.get("note") or ""])
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        finally:
            wb.close()

    # ------------------------------------------------------------------
    # 内部辅助
    # ------------------------------------------------------------------

    @staticmethod
    def _dedup_terms(terms: list[str]) -> list[str]:
        """去重保序（strip + 空过滤）。"""
        seen: set[str] = set()
        out: list[str] = []
        for term in terms or []:
            clean = str(term or "").strip()
            if not clean or clean in seen:
                continue
            seen.add(clean)
            out.append(clean)
        return out

    @staticmethod
    def _parse_header(header: tuple) -> dict[str, int]:
        """解析表头，返 ``{列名: 列索引}``。表头匹配 EXCEL_HEADERS。"""
        col_index: dict[str, int] = {}
        for idx, cell in enumerate(header):
            text = str(cell or "").strip()
            if text in EXCEL_HEADERS and text not in col_index:
                col_index[text] = idx
        # 必需列：卡片名称 + 关键词（风险等级/备注 可缺，缺则兜底）。
        if EXCEL_HEADER_CARD not in col_index or EXCEL_HEADER_TERM not in col_index:
            raise ValueError(
                "excel 表头需包含「卡片名称」和「关键词」列（规范：卡片名称,关键词,风险等级,备注）"
            )
        return col_index

    @staticmethod
    def _parse_row(row: tuple, col_index: dict[str, int]) -> Optional[dict] | object:
        """解析一行。

        返回值三态:
        * 正常 dict ``{name, term, risk_level, note}`` — 有效行。
        * ``None`` — 整行空或无 term，静默跳过（不计 rejected）。
        * ``_REJECTED_ROW`` 哨兵 — 风险等级非法值，记入 rejected。
        """
        def cell(idx: int) -> str:
            if idx >= len(row):
                return ""
            return str(row[idx] or "").strip()

        name = cell(col_index[EXCEL_HEADER_CARD])
        term = cell(col_index[EXCEL_HEADER_TERM])
        risk = cell(col_index[EXCEL_HEADER_RISK]) if EXCEL_HEADER_RISK in col_index else ""
        note = cell(col_index[EXCEL_HEADER_NOTE]) if EXCEL_HEADER_NOTE in col_index else ""

        # 整行空 → 跳过（不计 rejected）。
        if not name and not term and not risk and not note:
            return None
        # 无关键词的行视为噪音行跳过（不计 rejected）。
        if not term:
            return None

        # 风险等级合法值校验：高/中/低。非法值 → 该行 rejected。
        if risk and risk not in RISK_LEVELS:
            return _REJECTED_ROW

        # 风险等级缺省 → 兜底「中」。
        if not risk:
            risk = RISK_MEDIUM

        return {"name": name, "term": term, "risk_level": risk, "note": note or None}

    async def _merge_rows(
        self, db: AsyncSession, parsed_rows: list[dict], rejected: int
    ) -> dict:
        """合并追加去重：同名卡片存在则追加新词（已有跳过）+ 覆盖 risk/note；不存在则新建。

        去重维度为 ``(card_id, term)``：同一 excel 内同卡同词的重复行也按
        「已有跳过」处理（第二行记入 skipped，不重复插），否则命中
        ``uq_keyword_terms_card_id_term`` 唯一约束 → IntegrityError。
        """
        # 按卡片名分组（保序）：name → {risk_level(首行非空), note(首行非空), terms[]}.
        groups: dict[str, dict] = {}
        order: list[str] = []
        skipped_terms = 0  # 同一 excel 内同卡同词重复行 + DB 已有词的合计。
        for row in parsed_rows:
            name = row["name"]
            if name not in groups:
                groups[name] = {
                    "risk_level": row["risk_level"],
                    "note": row["note"],
                    "terms": [],
                    "seen_terms": set(),
                }
                order.append(name)
            else:
                # 首行非空值优先（已有则不覆盖）。
                grp = groups[name]
                if not grp["risk_level"] and row["risk_level"]:
                    grp["risk_level"] = row["risk_level"]
                if not grp["note"] and row["note"]:
                    grp["note"] = row["note"]
            # 同一 excel 内同卡同词去重（保序）：重复行不进 terms，由调用方
            # 统计 skipped——避免唯一约束冲突 + 符合「已有 term 跳过」语义。
            grp = groups[name]
            term = row["term"]
            if term in grp["seen_terms"]:
                skipped_terms += 1
                continue
            grp["seen_terms"].add(term)
            grp["terms"].append(term)

        # 查现有卡片（按 name 一次查）。
        existing_result = await db.execute(
            select(KeywordCard).where(KeywordCard.name.in_(order))
        )
        existing_by_name: dict[str, KeywordCard] = {
            c.name: c for c in existing_result.scalars().all()
        }

        created_cards = 0
        appended_cards = 0
        new_terms = 0

        for name in order:
            grp = groups[name]
            if name in existing_by_name:
                card = existing_by_name[name]
                # 覆盖 risk_level/note（用 excel 新值）。
                card.risk_level = grp["risk_level"] or card.risk_level
                if grp["note"] is not None:
                    card.note = grp["note"]
                appended_cards += 1
                # 追加新 term（已有跳过）。
                existing_terms_result = await db.execute(
                    select(KeywordTerm.term).where(KeywordTerm.card_id == card.id)
                )
                existing_terms = {t for t in existing_terms_result.scalars().all()}
                for term in grp["terms"]:
                    if term in existing_terms:
                        skipped_terms += 1
                        continue
                    db.add(KeywordTerm(card_id=card.id, term=term))
                    new_terms += 1
            else:
                # 新建卡片。
                card = KeywordCard(
                    name=name,
                    risk_level=grp["risk_level"] or RISK_MEDIUM,
                    note=grp["note"],
                )
                db.add(card)
                await db.flush()
                created_cards += 1
                for term in grp["terms"]:
                    db.add(KeywordTerm(card_id=card.id, term=term))
                    new_terms += 1

        await db.flush()
        return {
            "created_cards": created_cards,
            "appended_cards": appended_cards,
            "new_terms": new_terms,
            "skipped_terms": skipped_terms,
            "rejected_rows": rejected,
        }
