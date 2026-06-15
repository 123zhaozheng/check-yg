# -*- coding: utf-8 -*-
"""Excel parser using openpyxl."""

import logging
from pathlib import Path
from typing import Any, List, Optional

import openpyxl

from .base import BaseParser, RawTable

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parser for Excel files using openpyxl."""

    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']

    @staticmethod
    def _cell_to_str(cell: Any) -> str:
        """Convert cell value to string."""
        if cell is None:
            return ""
        return str(cell).strip()

    @staticmethod
    def _escape_html(text: str) -> str:
        """转义 HTML 特殊字符."""
        return (text
                .replace('&', '&amp;')
                .replace('<', '&lt;')
                .replace('>', '&gt;')
                .replace('"', '&quot;'))

    def extract_raw_tables(self, file_path: Path) -> List[RawTable]:
        """从Excel文件中提取原始表格数据."""
        if not self.can_parse(file_path):
            return []

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            tables = []

            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                raw_table = self._parse_sheet_to_raw(ws, sheet_idx, sheet_name)
                if raw_table and raw_table.rows:
                    tables.append(raw_table)

            wb.close()
            return tables

        except Exception as e:
            self.logger.error("Failed to extract raw tables from Excel %s: %s",
                            file_path.name, e)
            return []

    def extract_non_table_context(self, file_path: Path, max_chars: int = 2000) -> str:
        """Excel files are primarily tabular; return empty string."""
        return ""

    def _parse_sheet_to_raw(self, ws, table_index: int, sheet_name: str) -> Optional[RawTable]:
        """将工作表解析为原始表格数据."""
        rows = []

        for row in ws.iter_rows(values_only=True):
            cells = [self._cell_to_str(cell) for cell in row]

            # 跳过空行
            if not any(cells):
                continue

            rows.append(cells)

        if not rows:
            return None

        # 生成HTML格式的原始内容
        html_rows = []
        for row in rows:
            td_elements = ''.join(f'<td>{self._escape_html(cell)}</td>' for cell in row)
            html_rows.append(f'<tr>{td_elements}</tr>')

        html_content = f'<table data-sheet="{sheet_name}">\n' + '\n'.join(html_rows) + '\n</table>'

        return RawTable(
            table_index=table_index,
            html_content=html_content,
            rows=rows
        )
