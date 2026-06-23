# -*- coding: utf-8 -*-
"""06-23-tab keyword library tests: CRUD + admin 鉴权 + 409 + excel 合并去重.

覆盖（prd §7）:
* 卡片 CRUD（list/create/update/delete + detail）。
* admin 鉴权：非 admin 调 POST/PUT/DELETE/import → 403；list/export/detail GET ok。
* 删已引用卡片返 409（命中引用存在时）。
* excel 导入合并追加去重：同名卡片追加新词（已有跳过）+ risk/note 覆盖；新建卡片。
* excel 导入返统计（新建/追加/新增词/跳过/拒绝）。
* excel 导出：一行一词，卡片名连续多行。
* excel 导入非法风险等级行跳过并记入 rejected。

sqlite 内存 DB + openpyxl 生成测试 xlsx — no real network.
"""

import io

import pytest
from openpyxl import Workbook
from sqlalchemy import select

from app.auth.dependencies import get_current_user
from app.main import app
from app.models import KeywordCard, KeywordHit, KeywordTerm, Role, User


# ---------------------------------------------------------------------------
# Admin-user client helper（admin 鉴权测试需要 admin 角色）
# ---------------------------------------------------------------------------


async def _admin_client(db_session, client):
    """Override get_current_user to return an admin user for the duration of the block."""
    session, auditor = db_session
    admin_role = Role(name="admin")
    admin_user = User(
        username="admin-test",
        email="admin@example.com",
        hashed_password="x",
        role=admin_role,
        is_active=True,
    )
    session.add_all([admin_role, admin_user])
    await session.commit()
    await session.refresh(admin_user)

    async def override_current_user():
        return admin_user

    app.dependency_overrides[get_current_user] = override_current_user
    yield client

    async def restore_current_user():
        return auditor

    app.dependency_overrides[get_current_user] = restore_current_user


def _make_xlsx(rows: list[list]) -> bytes:
    """Build an xlsx with header 卡片名称,关键词,风险等级,备注 + the given rows."""
    wb = Workbook()
    try:
        ws = wb.active
        ws.append(["卡片名称", "关键词", "风险等级", "备注"])
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 卡片 CRUD
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_create_card_admin(db_session, client):
    """POST /api/keyword-library/cards (admin) creates a card with terms."""
    async for c in _admin_client(db_session, client):
        resp = await c.post(
            "/api/keyword-library/cards",
            json={
                "name": "高风险对手",
                "risk_level": "高",
                "note": "敏感关键词集",
                "terms": ["张三", "李四", "张三"],  # 重复词去重
            },
        )
        assert resp.status_code == 201, resp.text
        body = resp.json()
        assert body["name"] == "高风险对手"
        assert body["risk_level"] == "高"
        assert body["note"] == "敏感关键词集"
        # 重复词去重 → 2 terms。
        assert len(body["terms"]) == 2
        assert {t["term"] for t in body["terms"]} == {"张三", "李四"}


@pytest.mark.asyncio
async def test_list_cards_returns_term_count(db_session, client):
    """GET /api/keyword-library/cards returns term_count per card (any login user)."""
    session, _user = db_session
    card = KeywordCard(name="卡A", risk_level="中", note="n1")
    session.add(card)
    await session.commit()
    await session.refresh(card)
    session.add_all(
        [
            KeywordTerm(card_id=card.id, term="词1"),
            KeywordTerm(card_id=card.id, term="词2"),
            KeywordTerm(card_id=card.id, term="词3"),
        ]
    )
    await session.commit()

    resp = await client.get("/api/keyword-library/cards")
    assert resp.status_code == 200
    items = resp.json()
    assert len(items) == 1
    assert items[0]["name"] == "卡A"
    assert items[0]["term_count"] == 3
    assert items[0]["risk_level"] == "中"


@pytest.mark.asyncio
async def test_get_card_detail(db_session, client):
    """GET /api/keyword-library/cards/{id} returns card + terms (any login user)."""
    session, _user = db_session
    card = KeywordCard(name="卡详情", risk_level="低")
    session.add(card)
    await session.commit()
    await session.refresh(card)
    session.add_all(
        [KeywordTerm(card_id=card.id, term="词X"), KeywordTerm(card_id=card.id, term="词Y")]
    )
    await session.commit()

    resp = await client.get(f"/api/keyword-library/cards/{card.id}")
    assert resp.status_code == 200
    body = resp.json()
    assert body["name"] == "卡详情"
    assert body["risk_level"] == "低"
    assert len(body["terms"]) == 2


@pytest.mark.asyncio
async def test_get_card_not_found_404(client):
    """GET nonexistent card → 404."""
    resp = await client.get("/api/keyword-library/cards/99999")
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_update_card_terms_full_replace(db_session, client):
    """PUT /api/keyword-library/cards/{id} (admin) — terms 全量替换 + name/risk/note 可改。"""
    session, _user = db_session
    card = KeywordCard(name="原卡", risk_level="中", note="原备注")
    session.add(card)
    await session.commit()
    await session.refresh(card)
    session.add_all(
        [KeywordTerm(card_id=card.id, term="旧词1"), KeywordTerm(card_id=card.id, term="旧词2")]
    )
    await session.commit()
    cid = card.id

    async for c in _admin_client(db_session, client):
        resp = await c.put(
            f"/api/keyword-library/cards/{cid}",
            json={
                "name": "新卡名",
                "risk_level": "高",
                "note": "新备注",
                "terms": ["新词A", "新词B"],
            },
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["name"] == "新卡名"
        assert body["risk_level"] == "高"
        assert body["note"] == "新备注"
        assert {t["term"] for t in body["terms"]} == {"新词A", "新词B"}

    # 旧词已删（全量替换）。
    terms = (
        await session.execute(select_terms_by_card(cid))
    ).scalars().all()
    assert {t.term for t in terms} == {"新词A", "新词B"}


def select_terms_by_card(card_id: int):
    from sqlalchemy import select

    return select(KeywordTerm).where(KeywordTerm.card_id == card_id)


@pytest.mark.asyncio
async def test_delete_card_cascades_terms(db_session, client):
    """DELETE /api/keyword-library/cards/{id} (admin) cascades terms."""
    session, _user = db_session
    card = KeywordCard(name="删卡", risk_level="低")
    session.add(card)
    await session.commit()
    await session.refresh(card)
    session.add_all(
        [KeywordTerm(card_id=card.id, term="词1"), KeywordTerm(card_id=card.id, term="词2")]
    )
    await session.commit()
    cid = card.id

    async for c in _admin_client(db_session, client):
        resp = await c.delete(f"/api/keyword-library/cards/{cid}")
        assert resp.status_code == 204

    # 卡片 + terms 都没了。
    from sqlalchemy import select

    gone_card = (
        await session.execute(select(KeywordCard).where(KeywordCard.id == cid))
    ).scalar_one_or_none()
    assert gone_card is None
    remaining_terms = (
        await session.execute(select(KeywordTerm).where(KeywordTerm.card_id == cid))
    ).scalars().all()
    assert remaining_terms == []


# ---------------------------------------------------------------------------
# admin 鉴权拒绝 auditor 改
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_non_admin_rejected_from_crud_and_import(client):
    """Non-admin (auditor) → POST/PUT/DELETE/import return 403; GET list/detail/export ok."""
    # GET list / export allowed for any login user.
    resp = await client.get("/api/keyword-library/cards")
    assert resp.status_code == 200
    resp = await client.get("/api/keyword-library/export")
    assert resp.status_code == 200

    # POST → 403.
    resp = await client.post(
        "/api/keyword-library/cards",
        json={"name": "x", "risk_level": "中", "terms": ["a"]},
    )
    assert resp.status_code == 403

    # PUT → 403.
    resp = await client.put(
        "/api/keyword-library/cards/1",
        json={"name": "y"},
    )
    assert resp.status_code == 403

    # DELETE → 403.
    resp = await client.delete("/api/keyword-library/cards/1")
    assert resp.status_code == 403

    # import → 403.
    xlsx = _make_xlsx([["卡", "词", "中", ""]])
    resp = await client.post(
        "/api/keyword-library/import",
        files={"file": ("kw.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 403


# ---------------------------------------------------------------------------
# 删已引用卡片返 409
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_delete_card_with_hits_rejected_409(db_session, client):
    """DELETE a card referenced by keyword_hits → 409; without hits → 204."""
    from app.models import FlowRecordRow, Task

    session, _user = db_session
    task = Task(title="t", owner_id=_user.id, status="completed")
    session.add(task)
    await session.commit()
    await session.refresh(task)
    record = FlowRecordRow(
        task_id=task.id,
        record_type="standard",
        row_index=1,
        is_valid=True,
        counterparty_name="张三",
    )
    session.add(record)
    await session.commit()
    await session.refresh(record)
    card = KeywordCard(name="引用卡", risk_level="高")
    session.add(card)
    await session.commit()
    await session.refresh(card)
    term = KeywordTerm(card_id=card.id, term="张三")
    session.add(term)
    await session.commit()
    await session.refresh(term)
    session.add(
        KeywordHit(
            task_id=task.id,
            flow_record_id=record.id,
            keyword_card_id=card.id,
            keyword_term_id=term.id,
            match_type="精确匹配",
            confidence=100,
            risk_level="高",
            matched_field="counterparty_name",
            matched_snippet="张三",
            status="pending",
        )
    )
    await session.commit()
    cid = card.id

    async for c in _admin_client(db_session, client):
        resp = await c.delete(f"/api/keyword-library/cards/{cid}")
        assert resp.status_code == 409
        assert "引用" in resp.json()["detail"]


# ---------------------------------------------------------------------------
# excel 导入合并追加去重 + 统计 + 拒绝行
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_import_excel_merges_same_name_card(db_session, client):
    """同名卡片存在 → 追加新词（已有跳过）+ risk/note 覆盖；返统计。"""
    session, _user = db_session
    card = KeywordCard(name="已有卡", risk_level="低", note="旧备注")
    session.add(card)
    await session.commit()
    await session.refresh(card)
    session.add_all(
        [KeywordTerm(card_id=card.id, term="旧词"), KeywordTerm(card_id=card.id, term="共有")]
    )
    await session.commit()
    cid = card.id

    xlsx = _make_xlsx(
        [
            ["已有卡", "共有", "高", "新备注"],  # 共有 已存在 → 跳过
            ["已有卡", "新词1", "高", "新备注"],  # 新词
            ["已有卡", "新词2", "高", "新备注"],  # 新词
        ]
    )

    async for c in _admin_client(db_session, client):
        resp = await c.post(
            "/api/keyword-library/import",
            files={"file": ("kw.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200, resp.text
        stats = resp.json()
        assert stats["created_cards"] == 0
        assert stats["appended_cards"] == 1
        assert stats["new_terms"] == 2  # 新词1 + 新词2
        assert stats["skipped_terms"] == 1  # 共有
        assert stats["rejected_rows"] == 0

    # 验证覆盖：risk_level 高 / note 新备注。
    await session.refresh(card)
    assert card.risk_level == "高"
    assert card.note == "新备注"
    # 验证 terms：旧词 + 共有 + 新词1 + 新词2 = 4。
    terms = (
        await session.execute(select_terms_by_card(cid))
    ).scalars().all()
    assert {t.term for t in terms} == {"旧词", "共有", "新词1", "新词2"}


@pytest.mark.asyncio
async def test_import_excel_creates_new_card(db_session, client):
    """同名卡片不存在 → 新建卡片 + terms。"""
    xlsx = _make_xlsx(
        [
            ["新卡A", "词1", "中", "备注A"],
            ["新卡A", "词2", "中", "备注A"],
            ["新卡B", "词3", "高", ""],
        ]
    )

    async for c in _admin_client(db_session, client):
        resp = await c.post(
            "/api/keyword-library/import",
            files={"file": ("kw.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200, resp.text
        stats = resp.json()
        assert stats["created_cards"] == 2
        assert stats["appended_cards"] == 0
        assert stats["new_terms"] == 3
        assert stats["skipped_terms"] == 0


@pytest.mark.asyncio
async def test_import_excel_dedups_within_file_duplicate_terms(db_session, client):
    """同一 excel 内同卡同词重复行 → 第二行按「已有跳过」记入 skipped，不重复插。

    回归：修复前 ``_merge_rows`` 把同卡同词两行都 append 到 grp["terms"]，
    新建卡片分支两词都插 → 命中 ``uq_keyword_terms_card_id_term`` 唯一约束
    → IntegrityError 500。去重维度必须是 ``(card_id, term)``，不能只查 DB 已有。
    """
    session, _user = db_session
    xlsx = _make_xlsx(
        [
            ["卡A", "词1", "高", "n"],  # 首次
            ["卡A", "词1", "高", "n"],  # 同 excel 内重复 → skipped
            ["卡A", "词2", "高", "n"],  # 新词
        ]
    )

    async for c in _admin_client(db_session, client):
        resp = await c.post(
            "/api/keyword-library/import",
            files={"file": ("kw.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200, resp.text
        stats = resp.json()
        assert stats["created_cards"] == 1
        assert stats["new_terms"] == 2  # 词1 + 词2
        assert stats["skipped_terms"] == 1  # 重复的词1

    # 验证 DB 只存 2 个唯一词（不命中唯一约束）。
    card = (
        await session.execute(select(KeywordCard).where(KeywordCard.name == "卡A"))
    ).scalar_one()
    terms = (
        await session.execute(select(KeywordTerm.term).where(KeywordTerm.card_id == card.id))
    ).scalars().all()
    assert set(terms) == {"词1", "词2"}


@pytest.mark.asyncio
async def test_import_excel_rejects_invalid_risk_level(db_session, client):
    """风险等级非法值（非法值→该行报错或兜底为「中」——本实现取「跳过该行并记入 rejected」）。"""
    xlsx = _make_xlsx(
        [
            ["卡X", "合法词", "高", ""],
            ["卡X", "非法词", "极高", ""],  # 风险等级「极高」非法 → rejected
            ["卡X", "缺省词", "", ""],  # 风险等级缺省 → 兜底「中」
        ]
    )

    async for c in _admin_client(db_session, client):
        resp = await c.post(
            "/api/keyword-library/import",
            files={"file": ("kw.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200, resp.text
        stats = resp.json()
        assert stats["rejected_rows"] == 1  # 极高
        # 缺省词兜底「中」仍入库；合法词入库 → 2 新词。
        assert stats["new_terms"] == 2


@pytest.mark.asyncio
async def test_import_excel_bad_header_422(client, db_session):
    """表头缺「关键词」列 → 422。"""
    wb = Workbook()
    try:
        ws = wb.active
        ws.append(["卡片名称", "备注"])  # 缺 关键词 / 风险等级
        ws.append(["卡", "n"])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx = buf.getvalue()
    finally:
        wb.close()

    async for c in _admin_client(db_session, client):
        resp = await c.post(
            "/api/keyword-library/import",
            files={"file": ("kw.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 422


# ---------------------------------------------------------------------------
# excel 导出
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_excel_one_row_per_term(db_session, client):
    """导出：一行一词，卡片名连续多行，表头 卡片名称,关键词,风险等级,备注。"""
    session, _user = db_session
    card1 = KeywordCard(name="卡1", risk_level="高", note="n1")
    card2 = KeywordCard(name="卡2", risk_level="低", note=None)
    session.add_all([card1, card2])
    await session.commit()
    await session.refresh(card1)
    await session.refresh(card2)
    session.add_all(
        [
            KeywordTerm(card_id=card1.id, term="词A"),
            KeywordTerm(card_id=card1.id, term="词B"),
            KeywordTerm(card_id=card2.id, term="词C"),
        ]
    )
    await session.commit()

    resp = await client.get("/api/keyword-library/export")
    assert resp.status_code == 200
    # 解析返的 xlsx。
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    # 表头 + 3 词行。
    assert rows[0] == ("卡片名称", "关键词", "风险等级", "备注")
    bodies = rows[1:]
    # 卡1 两行（词A/词B）+ 卡2 一行（词C）。
    assert ("卡1", "词A", "高", "n1") in bodies
    assert ("卡1", "词B", "高", "n1") in bodies
    assert ("卡2", "词C", "低", None) in bodies
