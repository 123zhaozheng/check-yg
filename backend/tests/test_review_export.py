# -*- coding: utf-8 -*-
"""Review/report/export service tests."""

import zipfile

import openpyxl
import pytest

from app.core.matcher import MatchType, NameMatcher
from app.auth.permissions import check_task_permission
from app.models import CustomerList, CustomerListItem, Document, Role, Task, User
from app.services.review_service import ReviewService


def test_matcher_priority_exact_masked_fuzzy():
    matcher = NameMatcher(fuzzy_threshold=0.5)

    exact = matcher.match("张三", "付款给张三")
    assert exact is not None
    assert exact.match_type == MatchType.EXACT

    masked = matcher.match("张三", "付款给张*")
    assert masked is not None
    assert masked.match_type == MatchType.MASKED

    fuzzy = matcher.match("张三丰", "付款给张三峰")
    assert fuzzy is not None
    assert fuzzy.match_type == MatchType.FUZZY


@pytest.mark.asyncio
async def test_review_service_persists_matches(db_session):
    session, user = db_session
    task, customer_list = await _seed_review_data(session, user.id)

    review = await ReviewService().run_review(
        session,
        task_id=task.id,
        customer_list_id=customer_list.id,
        match_config={"fuzzy_threshold": 0.5},
    )
    matches, total = await ReviewService().list_matches(session, review.id)

    assert review.status == "completed"
    assert total == 3
    assert [match.match_type for match in matches] == ["exact", "masked", "fuzzy"]
    assert matches[0].counterparty_name == "张三"


@pytest.mark.asyncio
async def test_review_api_denies_unauthorized_task(client, db_session):
    session, _user = db_session
    other = User(
        username="other",
        email="other@example.com",
        hashed_password="x",
        role_id=1,
        is_active=True,
    )
    task = Task(title="other task", owner=other, status="completed")
    customer_list = CustomerList(name="名单", owner=other, row_count=1)
    customer_list.items.append(CustomerListItem(name="张三"))
    session.add_all([other, task, customer_list])
    await session.commit()
    await session.refresh(task)
    await session.refresh(customer_list)

    response = await client.post(
        f"/api/tasks/{task.id}/review",
        json={"customer_list_id": customer_list.id},
    )

    assert response.status_code == 403


@pytest.mark.asyncio
async def test_admin_has_task_write_permission(db_session):
    session, _user = db_session
    admin_role = Role(name="admin")
    admin = User(
        username="admin-user",
        email="admin-user@example.com",
        hashed_password="x",
        role=admin_role,
        is_active=True,
    )
    owner = User(
        username="task-owner",
        email="task-owner@example.com",
        hashed_password="x",
        role_id=1,
        is_active=True,
    )
    task = Task(title="owned task", owner=owner, status="completed")
    session.add_all([admin_role, admin, owner, task])
    await session.commit()
    await session.refresh(admin)
    await session.refresh(task)

    assert await check_task_permission(session, admin, task.id, required_role="write")


@pytest.mark.asyncio
async def test_review_report_and_exports_are_downloadable(client, db_session, temp_output_dir):
    session, user = db_session
    task, customer_list = await _seed_review_data(session, user.id)

    review_response = await client.post(
        f"/api/tasks/{task.id}/review",
        json={"customer_list_id": customer_list.id, "match_config": {"fuzzy_threshold": 0.5}},
    )
    assert review_response.status_code == 200
    review_id = review_response.json()["id"]

    matches_response = await client.get(f"/api/reviews/{review_id}/matches?page=1&page_size=2")
    assert matches_response.status_code == 200
    assert matches_response.json()["total"] == 3
    assert len(matches_response.json()["items"]) == 2

    report_response = await client.post(
        f"/api/tasks/{task.id}/report",
        json={"review_id": review_id},
    )
    assert report_response.status_code == 200
    report_id = report_response.json()["id"]
    assert "审计报告" in report_response.json()["content"]

    report_download = await client.get(f"/api/reports/{report_id}/download")
    assert report_download.status_code == 200

    excel_response = await client.post(
        f"/api/tasks/{task.id}/export/excel",
        json={"review_id": review_id},
    )
    assert excel_response.status_code == 200
    excel_path = excel_response.json()["file_path"]
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        assert "标准化流水" in workbook.sheetnames
        assert "匹配详情" in workbook.sheetnames
    finally:
        workbook.close()

    bundle_response = await client.post(
        f"/api/tasks/{task.id}/export/bundle",
        json={"review_id": review_id},
    )
    assert bundle_response.status_code == 200
    bundle_path = bundle_response.json()["file_path"]
    with zipfile.ZipFile(bundle_path) as zf:
        assert "skill_manifest.json" in zf.namelist()
        assert "current_task/review_result.json" in zf.namelist()

    export_download = await client.get(f"/api/exports/{bundle_response.json()['id']}/download")
    assert export_download.status_code == 200


async def _seed_review_data(session, owner_id: int):
    task = Task(title="审查任务", owner_id=owner_id, status="completed")
    task.documents.append(
        Document(
            filename="flow.xlsx",
            original_path="flow.xlsx",
            status="completed",
            flow_tables={
                "records": [
                    {
                        "source_file": "flow.xlsx",
                        "original_row": 2,
                        "transaction_time": "2026-01-01 10:00:00",
                        "counterparty_name": "张三",
                        "counterparty_account": "1001",
                        "amount": "100.00",
                        "summary": "exact",
                    },
                    {
                        "source_file": "flow.xlsx",
                        "original_row": 3,
                        "transaction_time": "2026-01-01 11:00:00",
                        "counterparty_name": "李*",
                        "counterparty_account": "1002",
                        "amount": "200.00",
                        "summary": "masked",
                    },
                    {
                        "source_file": "flow.xlsx",
                        "original_row": 4,
                        "transaction_time": "2026-01-01 12:00:00",
                        "counterparty_name": "王小名",
                        "counterparty_account": "1003",
                        "amount": "300.00",
                        "summary": "fuzzy",
                    },
                ]
            },
        )
    )
    customer_list = CustomerList(name="客户名单", owner_id=owner_id, row_count=3)
    customer_list.items.extend(
        [
            CustomerListItem(name="张三"),
            CustomerListItem(name="李四"),
            CustomerListItem(name="王小明"),
        ]
    )
    session.add_all([task, customer_list])
    await session.commit()
    await session.refresh(task)
    await session.refresh(customer_list)
    return task, customer_list
