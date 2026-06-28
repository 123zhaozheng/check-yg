# -*- coding: utf-8 -*-
"""S8 导出+设置+辅助页闭环 API tests.

Covers (per prd §验证 新增后端单测):
* 报告导出 3 格式（pdf/docx/html 产物生成 + 文件存在 + content-type）.
* 数据导出 6 组合（3 scope × 2 format）.
* 导出历史列表 + 预览取样.
* change-password（旧密码错 / 成功 / 新密码过短）.
* users/me PATCH（改 username / 冲突 409）.
* settings/schema 返回元数据.
"""

from pathlib import Path

import openpyxl
import pytest

from app.auth.password import hash_password, verify_password
from app.models import (
    ExportFile,
    FlowRecordRow,
    Finding,
    Report,
    ReportChapter,
    Task,
    User,
)


# ---------------------------------------------------------------------------
# Seed helpers
# ---------------------------------------------------------------------------


async def _seed_task_with_report(session, user_id: int) -> tuple[Task, Report]:
    """Seed a completed task with a chaptered report (8 chapters)."""
    task = Task(
        title="导出测试任务",
        owner_id=user_id,
        status="completed",
        config={"cleaning_committed": "2026-06-20 10:00"},
    )
    session.add(task)
    await session.commit()
    await session.refresh(task)

    report = Report(
        task_id=task.id,
        review_id=None,
        format="markdown",
        content_path=f"report_chapters/task_{task.id}",
        status="draft",
    )
    session.add(report)
    await session.commit()
    await session.refresh(report)

    titles = [
        "概述", "被审查对象", "数据范围", "完整性校验（余额）",
        "关键词审查", "异常发现汇总", "风险评估", "结论建议",
    ]
    for idx, title in enumerate(titles):
        session.add(
            ReportChapter(
                report_id=report.id,
                title=title,
                content=f"## {title}\n\n章节 {idx} 正文内容。",
                order_index=idx,
            )
        )
    await session.commit()
    return task, report


async def _seed_flow_records(session, task_id: int) -> None:
    """Seed 2 standard + 1 excluded flow_records."""
    session.add_all([
        FlowRecordRow(
            task_id=task_id,
            record_type="standard",
            row_index=1,
            is_valid=True,
            transaction_time="2026-01-01 10:00:00",
            counterparty_name="张三",
            counterparty_account="1001",
            amount="100.00",
            raw_amount="100.00",
            summary="测试",
            transaction_type="收入",
            raw_payload={"cells": ["2026-01-01", "张三", "100"]},
        ),
        FlowRecordRow(
            task_id=task_id,
            record_type="standard",
            row_index=2,
            is_valid=True,
            transaction_time="2026-01-02 11:00:00",
            counterparty_name="李四",
            counterparty_account="1002",
            amount="200.00",
            raw_amount="-200.00",
            summary="测试2",
            transaction_type="支出",
            raw_payload={"cells": ["2026-01-02", "李四", "200"]},
        ),
        FlowRecordRow(
            task_id=task_id,
            record_type="excluded",
            row_index=3,
            is_valid=False,
            exclude_reason="classifier: not flow table",
            raw_payload={"cells": ["合计", "300"]},
        ),
    ])
    await session.commit()


async def _seed_finding(session, task_id: int) -> Finding:
    finding = Finding(
        task_id=task_id,
        type="大额交易",
        severity="high",
        description="单笔金额超阈值",
        counterparty="张三",
        amount="100000.00",
        confidence=0.9,
        status="pending",
    )
    session.add(finding)
    await session.commit()
    await session.refresh(finding)
    return finding


# ---------------------------------------------------------------------------
# 报告导出 3 格式
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_report_pdf_generates_file(client, db_session, temp_output_dir):
    """POST /tasks/{id}/export/report format=pdf 生成 pdf 文件."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)

    resp = await client.post(
        f"/api/tasks/{task.id}/export/report",
        json={"format": "pdf", "include_annotations": False},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["format"] == "pdf"
    assert body["scope"] == "report"
    path = Path(body["file_path"])
    assert path.exists()
    # PDF magic bytes.
    assert path.read_bytes()[:4] == b"%PDF"


@pytest.mark.asyncio
async def test_export_report_docx_generates_file(client, db_session, temp_output_dir):
    """POST /tasks/{id}/export/report format=docx 生成 docx 文件."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)

    resp = await client.post(
        f"/api/tasks/{task.id}/export/report",
        json={"format": "docx", "include_annotations": False},
    )
    assert resp.status_code == 200
    assert resp.json()["format"] == "docx"
    assert resp.json()["scope"] == "report"
    path = Path(resp.json()["file_path"])
    assert path.exists()
    # docx is a zip; magic bytes PK.
    assert path.read_bytes()[:2] == b"PK"


@pytest.mark.asyncio
async def test_export_report_html_generates_file(client, db_session, temp_output_dir):
    """POST /tasks/{id}/export/report format=html 生成自包含 html."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)

    resp = await client.post(
        f"/api/tasks/{task.id}/export/report",
        json={"format": "html", "include_annotations": True},
    )
    assert resp.status_code == 200
    assert resp.json()["format"] == "html"
    path = Path(resp.json()["file_path"])
    assert path.exists()
    content = path.read_text(encoding="utf-8")
    assert "<html" in content
    assert task.title in content
    # 8 章标题都应出现.
    for title in [
        "概述", "被审查对象", "数据范围", "完整性校验（余额）",
        "关键词审查", "异常发现汇总", "风险评估", "结论建议",
    ]:
        assert title in content


@pytest.mark.asyncio
async def test_export_report_404_when_no_report(client, db_session, temp_output_dir):
    """报告不存在 → 404."""
    session, user = db_session
    task = Task(title="无报告任务", owner_id=user.id, status="completed")
    session.add(task)
    await session.commit()
    await session.refresh(task)

    resp = await client.post(
        f"/api/tasks/{task.id}/export/report",
        json={"format": "pdf", "include_annotations": False},
    )
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_export_report_409_when_report_generating(client, db_session, temp_output_dir):
    """报告仍在 generating 时拒绝导出，避免产出空章节文件."""
    session, user = db_session
    task, report = await _seed_task_with_report(session, user.id)
    report.status = "generating"
    await session.commit()

    resp = await client.post(
        f"/api/tasks/{task.id}/export/report",
        json={"format": "pdf", "include_annotations": False},
    )
    assert resp.status_code == 409


# ---------------------------------------------------------------------------
# 数据导出 6 组合（3 scope × 2 format）
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_data_standard_excel(client, db_session, temp_output_dir):
    """scope=standard format=excel 生成只含 standard 记录的 xlsx."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)
    await _seed_flow_records(session, task.id)

    resp = await client.post(
        f"/api/tasks/{task.id}/export/data",
        json={"scope": "standard", "format": "excel"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["scope"] == "standard"
    assert body["format"] == "excel"
    wb = openpyxl.load_workbook(Path(body["file_path"]), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # 1 header + 2 standard records.
        assert len(rows) == 3
        assert rows[0][0] == "记录ID"
    finally:
        wb.close()


@pytest.mark.asyncio
async def test_export_data_raw_csv(client, db_session, temp_output_dir):
    """scope=raw format=csv 生成含 raw_payload 的 csv（UTF-8 BOM）."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)
    await _seed_flow_records(session, task.id)

    resp = await client.post(
        f"/api/tasks/{task.id}/export/data",
        json={"scope": "raw", "format": "csv"},
    )
    assert resp.status_code == 200
    assert resp.json()["scope"] == "raw"
    assert resp.json()["format"] == "csv"
    path = Path(resp.json()["file_path"])
    raw = path.read_bytes()
    # UTF-8 BOM.
    assert raw[:3] == b"\xef\xbb\xbf"
    text = raw.decode("utf-8-sig")
    # raw 含 3 条记录（2 standard + 1 excluded）.
    lines = [l for l in text.splitlines() if l]
    assert len(lines) == 4  # 1 header + 3 records
    # raw_payload 列存在（表头含"原始载荷"）.
    assert "原始载荷" in lines[0]


@pytest.mark.asyncio
async def test_export_data_findings_excel(client, db_session, temp_output_dir):
    """scope=findings format=excel 生成 findings xlsx."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)
    await _seed_finding(session, task.id)

    resp = await client.post(
        f"/api/tasks/{task.id}/export/data",
        json={"scope": "findings", "format": "excel"},
    )
    assert resp.status_code == 200
    assert resp.json()["scope"] == "findings"
    wb = openpyxl.load_workbook(Path(resp.json()["file_path"]), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # 1 header + 1 finding.
        assert len(rows) == 2
        assert rows[0][0] == "发现ID"
        assert rows[1][1] == "大额交易"
    finally:
        wb.close()


@pytest.mark.asyncio
async def test_export_data_findings_csv(client, db_session, temp_output_dir):
    """scope=findings format=csv."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)
    await _seed_finding(session, task.id)

    resp = await client.post(
        f"/api/tasks/{task.id}/export/data",
        json={"scope": "findings", "format": "csv"},
    )
    assert resp.status_code == 200
    path = Path(resp.json()["file_path"])
    assert path.read_bytes()[:3] == b"\xef\xbb\xbf"


@pytest.mark.asyncio
async def test_export_data_standard_csv(client, db_session, temp_output_dir):
    """scope=standard format=csv."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)
    await _seed_flow_records(session, task.id)

    resp = await client.post(
        f"/api/tasks/{task.id}/export/data",
        json={"scope": "standard", "format": "csv"},
    )
    assert resp.status_code == 200
    path = Path(resp.json()["file_path"])
    text = path.read_text(encoding="utf-8-sig")
    lines = [l for l in text.splitlines() if l]
    # 1 header + 2 standard records.
    assert len(lines) == 3


@pytest.mark.asyncio
async def test_export_data_raw_excel(client, db_session, temp_output_dir):
    """scope=raw format=excel."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)
    await _seed_flow_records(session, task.id)

    resp = await client.post(
        f"/api/tasks/{task.id}/export/data",
        json={"scope": "raw", "format": "excel"},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(Path(resp.json()["file_path"]), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # 1 header + 3 raw records.
        assert len(rows) == 4
        # raw 表头含"原始载荷"列.
        assert "原始载荷" in rows[0]
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 导出历史 + 预览
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_list_task_exports(client, db_session, temp_output_dir):
    """GET /tasks/{id}/exports 返导出历史（按 created_at 降序）."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)

    # 先生成两份导出.
    await client.post(
        f"/api/tasks/{task.id}/export/report",
        json={"format": "pdf", "include_annotations": False},
    )
    await client.post(
        f"/api/tasks/{task.id}/export/data",
        json={"scope": "standard", "format": "excel"},
    )

    resp = await client.get(f"/api/tasks/{task.id}/exports")
    assert resp.status_code == 200
    body = resp.json()
    assert len(body) >= 2
    # 两条 scope 分别为 report / standard.
    scopes = {e["scope"] for e in body}
    assert "report" in scopes
    assert "standard" in scopes


@pytest.mark.asyncio
async def test_preview_report_returns_chapters(client, db_session, temp_output_dir):
    """GET /tasks/{id}/export/preview?scope=report 返前 2 章 content + 批注数."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)

    resp = await client.get(
        f"/api/tasks/{task.id}/export/preview",
        params={"scope": "report"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["scope"] == "report"
    assert len(body["sample"]) == 2
    assert body["sample"][0]["title"] == "概述"
    assert body["annotation_count"] == 0


@pytest.mark.asyncio
async def test_preview_standard_returns_rows(client, db_session, temp_output_dir):
    """GET /tasks/{id}/export/preview?scope=standard 返前 20 行 JSON."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)
    await _seed_flow_records(session, task.id)

    resp = await client.get(
        f"/api/tasks/{task.id}/export/preview",
        params={"scope": "standard"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["scope"] == "standard"
    assert len(body["sample"]) == 2  # only 2 standard records seeded


@pytest.mark.asyncio
async def test_preview_findings_returns_rows(client, db_session, temp_output_dir):
    """GET /tasks/{id}/export/preview?scope=findings 返 findings 行."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)
    await _seed_finding(session, task.id)

    resp = await client.get(
        f"/api/tasks/{task.id}/export/preview",
        params={"scope": "findings"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["scope"] == "findings"
    assert len(body["sample"]) == 1
    assert body["sample"][0]["type"] == "大额交易"


@pytest.mark.asyncio
async def test_preview_report_404_when_no_report(client, db_session, temp_output_dir):
    """报告不存在时 preview report → 404."""
    session, user = db_session
    task = Task(title="无报告预览", owner_id=user.id, status="completed")
    session.add(task)
    await session.commit()
    await session.refresh(task)

    resp = await client.get(
        f"/api/tasks/{task.id}/export/preview",
        params={"scope": "report"},
    )
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# owner 校验
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_endpoints_owner_only_403(client, db_session, temp_output_dir):
    """非 owner 不能导出."""
    from app.models import Role

    session, user = db_session
    other_role = Role(name="export_other")
    other = User(
        username="export-other",
        email="export-other@example.com",
        hashed_password="x",
        role=other_role,
        is_active=True,
    )
    task, _report = await _seed_task_with_report(session, user.id)
    session.add_all([other_role, other])
    await session.commit()

    from app.auth.dependencies import get_current_user
    from app.main import app

    async def _other_user():
        return other

    app.dependency_overrides[get_current_user] = _other_user
    try:
        r = await client.post(
            f"/api/tasks/{task.id}/export/report",
            json={"format": "pdf", "include_annotations": False},
        )
        assert r.status_code == 403
        r = await client.get(f"/api/tasks/{task.id}/exports")
        assert r.status_code == 403
    finally:
        async def _owner():
            return user
        app.dependency_overrides[get_current_user] = _owner


@pytest.mark.asyncio
async def test_export_download_works_for_new_formats(client, db_session, temp_output_dir):
    """GET /api/exports/{id}/download 可下载 S8 新格式产物."""
    session, user = db_session
    task, _report = await _seed_task_with_report(session, user.id)

    created = await client.post(
        f"/api/tasks/{task.id}/export/report",
        json={"format": "pdf", "include_annotations": False},
    )
    export_id = created.json()["id"]

    resp = await client.get(f"/api/exports/{export_id}/download")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"


# ---------------------------------------------------------------------------
# change-password
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_change_password_success(client, db_session):
    """旧密码正确 + 新密码长度足够 → 成功更新."""
    session, user = db_session
    user.hashed_password = hash_password("oldpass123")
    session.add(user)
    await session.commit()
    await session.refresh(user)

    resp = await client.post(
        "/api/auth/change-password",
        json={"old_password": "oldpass123", "new_password": "newpass456"},
    )
    assert resp.status_code == 200
    assert resp.json()["ok"] is True

    # DB 真实更新.
    from sqlalchemy import select
    row = (
        await session.execute(select(User).where(User.id == user.id))
    ).scalar_one()
    assert verify_password("newpass456", row.hashed_password)


@pytest.mark.asyncio
async def test_change_password_wrong_old(client, db_session):
    """旧密码错 → 400."""
    session, user = db_session
    user.hashed_password = hash_password("oldpass123")
    session.add(user)
    await session.commit()

    resp = await client.post(
        "/api/auth/change-password",
        json={"old_password": "wrongold", "new_password": "newpass456"},
    )
    assert resp.status_code == 400
    assert "旧密码" in resp.json()["detail"]


@pytest.mark.asyncio
async def test_change_password_too_short(client, db_session):
    """新密码长度 < 8 → 422."""
    session, user = db_session
    user.hashed_password = hash_password("oldpass123")
    session.add(user)
    await session.commit()

    resp = await client.post(
        "/api/auth/change-password",
        json={"old_password": "oldpass123", "new_password": "short"},
    )
    assert resp.status_code == 422


# ---------------------------------------------------------------------------
# users/me PATCH
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_update_me_username(client, db_session):
    """PATCH /users/me 改 username（当前用户改自己，非 admin-only）."""
    session, user = db_session

    resp = await client.patch(
        "/api/users/me",
        json={"username": "owner-renamed"},
    )
    assert resp.status_code == 200
    assert resp.json()["username"] == "owner-renamed"
    assert resp.json()["id"] == user.id

    # DB 真实更新.
    from sqlalchemy import select
    row = (
        await session.execute(select(User).where(User.id == user.id))
    ).scalar_one()
    assert row.username == "owner-renamed"


@pytest.mark.asyncio
async def test_update_me_email(client, db_session):
    """PATCH /users/me 改 email."""
    session, user = db_session

    resp = await client.patch(
        "/api/users/me",
        json={"email": "new-email@example.com"},
    )
    assert resp.status_code == 200
    assert resp.json()["email"] == "new-email@example.com"


@pytest.mark.asyncio
async def test_update_me_username_conflict_409(client, db_session):
    """username 冲突 → 409."""
    session, user = db_session
    from app.models import Role
    other_role = Role(name="me-conflict")
    other = User(
        username="taken-name",
        email="taken@example.com",
        hashed_password="x",
        role=other_role,
        is_active=True,
    )
    session.add_all([other_role, other])
    await session.commit()

    resp = await client.patch(
        "/api/users/me",
        json={"username": "taken-name"},
    )
    assert resp.status_code == 409


# ---------------------------------------------------------------------------
# settings/schema
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_settings_schema_returns_metadata(client, db_session):
    """GET /api/settings/schema 返回设置项元数据列表."""
    resp = await client.get("/api/settings/schema")
    assert resp.status_code == 200
    body = resp.json()
    assert isinstance(body, list)
    assert len(body) > 0
    # 每项含 key/category/type/label/description/value.
    first = body[0]
    for field in ["key", "category", "type", "label", "description", "value"]:
        assert field in first
    # 应包含 S8 新增设置项.
    keys = {item["key"] for item in body}
    assert "audit.fuzzy_threshold" in keys
    assert "audit.default_analysis_mode" in keys
    assert "llm.temperature" in keys
    assert "channel.bank.enabled" in keys
    # select 类型应带 options.
    mode_item = next(item for item in body if item["key"] == "audit.default_analysis_mode")
    assert mode_item["type"] == "select"
    assert mode_item["options"] == ["quick", "deep"]
    # boolean 类型.
    bank_item = next(item for item in body if item["key"] == "channel.bank.enabled")
    assert bank_item["type"] == "boolean"


@pytest.mark.asyncio
async def test_settings_schema_reflects_saved_values(client, db_session):
    """schema 的 value 优先用 DB 已存值."""
    session, user = db_session
    from app.models import Setting
    session.add(
        Setting(
            key="audit.fuzzy_threshold",
            value="0.85",
            category="audit",
            updated_by=user.id,
        )
    )
    await session.commit()

    resp = await client.get("/api/settings/schema")
    body = resp.json()
    item = next(i for i in body if i["key"] == "audit.fuzzy_threshold")
    assert item["value"] == "0.85"
