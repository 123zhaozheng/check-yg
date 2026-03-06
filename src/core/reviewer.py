# -*- coding: utf-8 -*-
"""
Reviewer - 简化版流水审查器
只做精确匹配和脱敏匹配，不调用LLM
"""

import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import openpyxl
from ..core.matcher import NameMatcher, MatchResult, MatchType
from ..core.customer import CustomerManager
from ..core.review_history import ReviewHistoryManager
from ..config import get_config

logger = logging.getLogger(__name__)


@dataclass
class ReviewMatch:
    """审查匹配结果"""
    customer_name: str
    counterparty_name: str
    counterparty_account: str
    match_type: str  # "精确匹配" / "脱敏匹配"
    confidence: int
    source_file: str
    row_index: int = 0
    transaction_time: str = ""
    amount: str = ""
    summary: str = ""
    
    def to_dict(self) -> dict:
        """转换为字典"""
        return {
            'customer_name': self.customer_name,
            'counterparty_name': self.counterparty_name,
            'counterparty_account': self.counterparty_account,
            'match_type': self.match_type,
            'confidence': self.confidence,
            'source_file': self.source_file,
            'row_index': self.row_index,
            'transaction_time': self.transaction_time,
            'amount': self.amount,
            'summary': self.summary,
        }


@dataclass
class ReviewResult:
    """审查结果"""
    review_id: str
    review_time: str
    flow_excel_path: str
    customer_excel_path: str
    total_customers: int
    matched_customers: int
    total_matches: int
    total_amount: float
    matches: List[ReviewMatch] = field(default_factory=list)
    writeback_error: str = ""
    
    def __post_init__(self):
        if self.matches is None:
            self.matches = []
    
    @property
    def total_amount_formatted(self) -> str:
        """格式化总金额"""
        return f"¥{self.total_amount:,.2f}"
    
    def to_dict(self) -> dict:
        """转换为字典"""
        from dataclasses import asdict
        result = asdict(self)
        result['matches'] = [m.to_dict() for m in self.matches]
        return result


class Reviewer:
    """
    流水审查器（简化版）
    
    只做精确匹配和脱敏匹配，不调用LLM
    """
    
    def __init__(self, config=None):
        self.config = config or get_config()
        self.matcher = NameMatcher(fuzzy_threshold=self.config.fuzzy_threshold)
        self.customer_manager = CustomerManager()
        self.review_history = ReviewHistoryManager(self.config.config_dir / "reviews")
    
    def load_flows(self, excel_path: str) -> List[dict]:
        """
        加载流水Excel
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            List[dict]: 流水记录列表
        """
        try:
            path = Path(excel_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {excel_path}")
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active

            rows_iter = ws.iter_rows(values_only=True)
            headers = []
            for row in rows_iter:
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
                break

            if not any(headers):
                wb.close()
                raise ValueError("流水Excel缺少表头")

            records = []
            row_index = 1
            for row in rows_iter:
                row_index += 1
                record = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = row[idx] if idx < len(row) else ""
                    record[header] = "" if value is None else str(value).strip()
                if record:
                    record["_row_index"] = row_index
                    records.append(record)
            wb.close()
            logger.info("加载流水: %d 条", len(records))
            return records
        except Exception as e:
            logger.error("加载流水Excel失败: %s", e)
            raise
    
    def load_customers(self, excel_path: str) -> int:
        """
        加载客户名单
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            int: 客户数量
        """
        count = self.customer_manager.load_from_excel(excel_path)
        logger.info("加载客户: %d 个", count)
        return count
    
    def run_review(
        self,
        flow_excel_path: str,
        customer_excel_path: str = "",
        customers: Optional[List[str]] = None
    ) -> ReviewResult:
        """
        执行审查
        
        Args:
            flow_excel_path: 流水Excel路径
            customer_excel_path: 客户名单Excel路径
            
        Returns:
            ReviewResult: 审查结果
        """
        # 加载数据
        flows = self.load_flows(flow_excel_path)
        if customers is not None:
            self.customer_manager.load_from_list(customers)
            customer_count = self.customer_manager.count
        else:
            customer_count = self.load_customers(customer_excel_path)

        # 过滤金额异常的流水（继续审查但丢弃该条）
        valid_flows = []
        skipped_amount_rows = 0
        for flow in flows:
            parsed_amount = self._parse_amount(flow.get('金额', ''))
            if parsed_amount is None:
                skipped_amount_rows += 1
                continue
            flow['_parsed_amount'] = parsed_amount
            valid_flows.append(flow)

        if skipped_amount_rows:
            logger.warning("发现 %d 条金额无效的流水，已跳过", skipped_amount_rows)

        # 匹配
        matches: List[ReviewMatch] = []
        best_match_by_row: dict = {}
        for flow in valid_flows:
            counterparty = str(flow.get('交易对手名', ''))
            counterparty_account = str(flow.get('交易对手账号', ''))
            if not counterparty:
                continue
            
            # 尝试匹配每个客户
            for customer in self.customer_manager:
                # 精确匹配
                if self.config.enable_exact_match:
                    result = self.matcher.match_exact(customer, counterparty)
                    if result:
                        match = self._create_match(
                            customer, result, flow, counterparty, counterparty_account
                        )
                        matches.append(match)
                        self._update_best_match(best_match_by_row, match)
                        continue
                
                # 脱敏匹配
                if self.config.enable_desensitized_match:
                    result = self.matcher.match_desensitized(customer, counterparty)
                    if result:
                        match = self._create_match(
                            customer, result, flow, counterparty, counterparty_account
                        )
                        matches.append(match)
                        self._update_best_match(best_match_by_row, match)
                        continue

                # 模糊匹配
                if self.config.enable_fuzzy_match:
                    result = self.matcher.match_fuzzy(customer, counterparty)
                    if result:
                        match = self._create_match(
                            customer, result, flow, counterparty, counterparty_account
                        )
                        matches.append(match)
                        self._update_best_match(best_match_by_row, match)

        # 统计
        total_amount = sum(
            float(flow.get('_parsed_amount', 0.0)) for flow in valid_flows
        )
        matched_customers = set(m.customer_name for m in matches)
        
        result = ReviewResult(
            review_id=datetime.now().strftime("%Y%m%d_%H%M%S"),
            review_time=datetime.now().isoformat(),
            flow_excel_path=flow_excel_path,
            customer_excel_path=customer_excel_path,
            total_customers=customer_count,
            matched_customers=len(matched_customers),
            total_matches=len(matches),
            total_amount=total_amount,
            matches=matches
        )
        try:
            self._write_back_results(flow_excel_path, best_match_by_row, matches)
        except Exception as exc:
            result.writeback_error = str(exc)
            logger.warning("写回流水Excel失败: %s", exc)
        try:
            self.review_history.save_review_result(result, flow_excel_path)
        except Exception as exc:
            logger.warning("保存审查历史失败: %s", exc)

        return result

    def _parse_amount(self, amount_value: object) -> Optional[float]:
        """Parse amount string to float; return None if invalid."""
        if amount_value is None:
            return None
        amount_str = str(amount_value).strip()
        if not amount_str:
            return None
        clean = (
            amount_str.replace(',', '')
            .replace('￥', '')
            .replace('¥', '')
            .replace('元', '')
            .replace('+', '')
            .replace('-', '')
        )
        try:
            return abs(float(clean))
        except (ValueError, TypeError):
            return None
    
    def _create_match(
        self,
        customer_name: str,
        match_result: MatchResult,
        flow: dict,
        counterparty: str,
        counterparty_account: str
    ) -> ReviewMatch:
        """创建审查匹配对象"""
        return ReviewMatch(
            customer_name=customer_name,
            counterparty_name=counterparty,
            counterparty_account=counterparty_account,
            match_type=match_result.match_type.value,
            confidence=match_result.confidence,
            source_file=str(flow.get('来源文件', '')),
            row_index=int(flow.get('_row_index', 0) or 0),
            transaction_time=str(flow.get('交易时间', '')),
            amount=str(flow.get('金额', '')),
            summary=str(flow.get('摘要', '')),
        )

    def _update_best_match(self, best_map: dict, match: ReviewMatch) -> None:
        row_index = match.row_index
        if row_index <= 0:
            return
        existing = best_map.get(row_index)
        if not existing:
            best_map[row_index] = match
            return

        def rank(m: ReviewMatch) -> tuple:
            type_rank = {
                "精确匹配": 3,
                "脱敏匹配": 2,
                "模糊匹配": 1,
            }.get(m.match_type, 0)
            return (int(m.confidence), type_rank)

        if rank(match) > rank(existing):
            best_map[row_index] = match

    def _write_back_results(
        self,
        flow_excel_path: str,
        best_match_by_row: dict,
        matches: List[ReviewMatch]
    ) -> None:
        path = Path(flow_excel_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {flow_excel_path}")

        wb = openpyxl.load_workbook(path)
        try:
            ws = wb.active

            headers = []
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=1, column=col).value
                headers.append(str(value).strip() if value is not None else "")
            if not any(headers):
                raise ValueError("流水Excel缺少表头")

            user_col = self._find_header_column(headers, "匹配用户")
            conf_col = self._find_header_column(headers, "匹配度")

            if user_col is None:
                user_col = ws.max_column + 1
                ws.cell(row=1, column=user_col, value="匹配用户")
            if conf_col is None:
                conf_col = ws.max_column + 1
                ws.cell(row=1, column=conf_col, value="匹配度")

            # 清空原有匹配列，避免旧数据残留
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=user_col, value=None)
                ws.cell(row=row, column=conf_col, value=None)

            # 写入最佳匹配
            for row_index, match in best_match_by_row.items():
                if row_index <= 1:
                    continue
                ws.cell(row=row_index, column=user_col, value=match.customer_name)
                ws.cell(row=row_index, column=conf_col, value=int(match.confidence))

            # 写入匹配详情Sheet
            self._write_match_details_sheet(wb, matches)

            wb.save(path)
        finally:
            wb.close()

    @staticmethod
    def _find_header_column(headers: List[str], name: str) -> Optional[int]:
        for idx, header in enumerate(headers, 1):
            if header == name:
                return idx
        return None

    @staticmethod
    def _write_match_details_sheet(wb, matches: List[ReviewMatch]) -> None:
        sheet_name = "匹配详情"
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        headers = [
            "流水行号",
            "匹配用户",
            "匹配度",
            "匹配类型",
            "来源文件",
            "交易时间",
            "对手名",
            "对手账号",
            "金额",
            "摘要",
        ]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, match in enumerate(matches, 2):
            data = [
                match.row_index,
                match.customer_name,
                int(match.confidence),
                match.match_type,
                match.source_file,
                match.transaction_time,
                match.counterparty_name,
                match.counterparty_account,
                match.amount,
                match.summary,
            ]
            for col_idx, value in enumerate(data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
