# -*- coding: utf-8 -*-
"""
Flow Excel Exporter - 导出流水记录到Excel
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..parsers.base import FLOW_EXCEL_COLUMNS, FlowRecord
from ..config import get_config

if TYPE_CHECKING:
    from ..core.extraction_result import ExtractionResult

logger = logging.getLogger(__name__)

# 星期几中文映射（0=周一 ... 6=周日）
_WEEKDAY_NAMES = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


def _compute_weekday(time_str: str) -> str:
    """纯程序化计算星期几，严格基于原始交易时间字符串，不依赖LLM。"""
    if not time_str or not time_str.strip():
        return "未知"
    try:
        dt = datetime.strptime(time_str.strip()[:19], "%Y-%m-%d %H:%M:%S")
    except (ValueError, IndexError):
        try:
            dt = datetime.strptime(time_str.strip()[:10], "%Y-%m-%d")
        except (ValueError, IndexError):
            return "未知"
    return _WEEKDAY_NAMES[dt.weekday()]


def _compute_is_rest_day(time_str: str) -> str:
    """纯程序化计算是否休息日，使用chinesecalendar库，不依赖LLM。"""
    if not time_str or not time_str.strip():
        return "未知"
    try:
        dt = datetime.strptime(time_str.strip()[:19], "%Y-%m-%d %H:%M:%S")
    except (ValueError, IndexError):
        try:
            dt = datetime.strptime(time_str.strip()[:10], "%Y-%m-%d")
        except (ValueError, IndexError):
            return "未知"
    try:
        import chinese_calendar
        return "否" if chinese_calendar.is_workday(dt.date()) else "是"
    except (ImportError, NotImplementedError):
        # chinese_calendar 未安装或日期超出覆盖范围
        return "未知"


class FlowExporter:
    """
    流水Excel导出器

    导出路径：~/.check-yg/flows/
    固定列结构：来源文件、原始行号、交易时间、交易对手名、交易对手账号、金额、摘要、收支类型、星期几、是否休息日
    """
    
    # 样式定义
    HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 列宽度
    COLUMN_WIDTHS = {
        '来源文件': 30,
        '原始行号': 10,
        '交易时间': 20,
        '交易对手名': 20,
        '交易对手账号': 20,
        '金额': 15,
        '摘要': 30,
        '收支类型': 12,
        '星期几': 10,
        '是否休息日': 12,
    }
    
    def __init__(self, output_folder: Optional[Path] = None):
        """
        初始化导出器
        
        Args:
            output_folder: 输出目录，默认 ~/.check-yg/flows/
        """
        if output_folder:
            self.output_folder = Path(output_folder)
        else:
            config = get_config()
            self.output_folder = config.config_dir / 'flows'
        
        self.output_folder.mkdir(parents=True, exist_ok=True)
    
    def export(
        self,
        records: List[FlowRecord],
        task_id: str,
        filename: Optional[str] = None,
        extraction_result: Optional['ExtractionResult'] = None
    ) -> Path:
        """
        导出流水记录到Excel

        Args:
            records: 流水记录列表
            task_id: 任务ID
            filename: 输出文件名（自动生成如果为空）
            extraction_result: 提取结果（用于生成处理汇总Sheet）

        Returns:
            Path: 导出文件路径
        """
        if not records:
            logger.warning("没有流水记录可导出")
            raise ValueError("没有流水记录可导出")
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            # 创建第一个工作表
            ws = wb.create_sheet("流水明细", 0)
        ws.title = "流水明细"
        
        # 写入表头
        for col_idx, header in enumerate(FLOW_EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal='center')
            
            # 设置列宽
            width = self.COLUMN_WIDTHS.get(header, 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 写入数据
        for row_idx, record in enumerate(records, 2):
            row_data = record.to_list()
            # 追加星期几和是否休息日（纯程序化计算，不依赖LLM）
            row_data.append(_compute_weekday(record.transaction_time))
            row_data.append(_compute_is_rest_day(record.transaction_time))
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                cell.alignment = Alignment(vertical='center')
        
        # 冻结首行
        ws.freeze_panes = 'A2'

        # 写入处理汇总Sheet（如果有提取结果）
        if extraction_result is not None:
            self._write_summary_sheet(wb, extraction_result)

        # 生成文件名
        if not filename:
            filename = f"流水_{task_id}.xlsx"
        
        # 保存文件
        output_path = self.output_folder / filename
        wb.save(output_path)
        
        logger.info("导出流水到: %s (%d 条记录)",
                   output_path, len(records))

        return output_path

    def _write_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        extraction_result: 'ExtractionResult'
    ) -> None:
        """写入处理汇总Sheet页"""
        sheet_name = "处理汇总"
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        # 汇总表头
        summary_headers = ["文档名称", "标准化流水数", "状态", "失败原因"]
        summary_widths = [30, 15, 10, 40]
        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col_idx)].width = summary_widths[col_idx - 1]

        # 收集失败文档的错误原因
        failed_error_map: dict = {}
        for err in extraction_result.errors:
            doc = err.get("document", "")
            error_msg = err.get("error", "")
            if doc in extraction_result.failed_documents:
                failed_error_map.setdefault(doc, []).append(error_msg)

        # 合并所有文档：per_document_stats 中的成功文档 + failed_documents 中的失败文档
        all_docs: dict = {}
        # 成功文档
        for doc_name, stats in extraction_result.per_document_stats.items():
            all_docs[doc_name] = {
                "record_count": stats.get("record_count", 0),
                "status": "成功",
                "reason": "",
            }
        # 失败文档
        for doc_name in extraction_result.failed_documents:
            reasons = failed_error_map.get(doc_name, [])
            reason_str = "; ".join(reasons) if reasons else "处理失败"
            all_docs[doc_name] = {
                "record_count": 0,
                "status": "失败",
                "reason": reason_str,
            }

        # 写入数据行
        for row_idx, (doc_name, info) in enumerate(all_docs.items(), 2):
            row_data = [
                doc_name,
                info["record_count"],
                info["status"],
                info["reason"],
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                cell.alignment = Alignment(vertical='center')

        # 冻结首行
        ws.freeze_panes = 'A2'