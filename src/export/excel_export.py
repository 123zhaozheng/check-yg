# -*- coding: utf-8 -*-
"""
Excel export module for audit reports
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..core.auditor import AuditResult, AuditMatch

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export audit results to Excel format"""
    
    # Style definitions
    HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    
    HIGH_RISK_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    MEDIUM_RISK_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    LOW_RISK_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column definitions
    COLUMNS = [
        ("客户姓名", 15),
        ("匹配文本", 15),
        ("匹配类型", 12),
        ("置信度", 10),
        ("来源文件", 30),
        ("行号", 8),
        ("交易时间", 20),
        ("交易金额", 15),
        ("收支类型", 10),
        ("摘要", 30),
    ]
    
    def __init__(self, output_folder: Optional[Path] = None):
        """
        Initialize exporter
        
        Args:
            output_folder: Folder to save exported files
        """
        self.output_folder = output_folder or Path("./data/reports")
    
    def export(
        self,
        result: AuditResult,
        filename: Optional[str] = None,
        include_summary: bool = True
    ) -> Path:
        """
        Export audit result to Excel file
        
        Args:
            result: AuditResult to export
            filename: Output filename (auto-generated if not provided)
            include_summary: Whether to include summary sheet
            
        Returns:
            Path to exported file
        """
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create summary sheet
        if include_summary:
            self._create_summary_sheet(wb, result)
        
        # Create details sheet
        self._create_details_sheet(wb, result)
        
        # Remove default sheet if we created others
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
        
        # Generate filename
        if not filename:
            filename = f"审计报告_{result.audit_id}.xlsx"
        
        # Ensure output folder exists
        self.output_folder.mkdir(parents=True, exist_ok=True)
        output_path = self.output_folder / filename
        
        # Save workbook
        wb.save(output_path)
        logger.info("Exported audit report to %s", output_path)
        
        return output_path
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, result: AuditResult) -> None:
        """Create summary sheet with statistics"""
        ws = wb.create_sheet("审计概览", 0)
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "员工-客户金额往来审计报告"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Summary data
        summary_data = [
            ("审计编号", result.audit_id),
            ("审计时间", result.audit_time),
            ("客户名单", result.customer_file),
            ("文档目录", result.document_folder),
            ("", ""),
            ("客户总数", result.total_customers),
            ("命中客户数", result.matched_customers),
            ("匹配记录数", result.total_matches),
            ("涉及金额", f"¥{result.total_amount:,.2f}"),
            ("风险等级", result.risk_level),
        ]
        
        start_row = 3
        for i, (label, value) in enumerate(summary_data):
            row = start_row + i
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            
            if label:
                ws[f'A{row}'].font = Font(bold=True)
            
            # Highlight risk level
            if label == "风险等级":
                if "高" in str(value):
                    ws[f'B{row}'].fill = self.HIGH_RISK_FILL
                elif "中" in str(value):
                    ws[f'B{row}'].fill = self.MEDIUM_RISK_FILL
                else:
                    ws[f'B{row}'].fill = self.LOW_RISK_FILL
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
    
    def _create_details_sheet(self, wb: openpyxl.Workbook, result: AuditResult) -> None:
        """Create details sheet with all matches"""
        ws = wb.create_sheet("匹配明细")
        
        # Write headers
        for col, (header, width) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Write data rows
        for row_idx, match in enumerate(result.matches, 2):
            data = [
                match.customer_name,
                match.matched_text,
                match.confidence,
                f"{match.confidence_score}%",
                Path(match.source_file).name,
                match.row_index + 1,
                match.transaction_time,
                match.amount,
                match.transaction_type,
                match.summary,
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.BORDER
                cell.alignment = Alignment(vertical='center')
            
            # Highlight by confidence
            if match.confidence == "精确匹配":
                pass  # No highlight
            elif match.confidence == "脱敏匹配":
                for col in range(1, len(data) + 1):
                    ws.cell(row=row_idx, column=col).fill = self.MEDIUM_RISK_FILL
            else:
                for col in range(1, len(data) + 1):
                    ws.cell(row=row_idx, column=col).fill = self.HIGH_RISK_FILL
        
        # Freeze header row
        ws.freeze_panes = 'A2'
