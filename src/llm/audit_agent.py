# -*- coding: utf-8 -*-
"""
LLM-backed audit report and QA helper.
"""

import json
import logging
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import requests

from ..config import get_config

logger = logging.getLogger(__name__)


REPORT_SYSTEM_PROMPT = """你是一个资深审计报告撰写助手。

你的任务是基于当前审查任务的结构化数据，输出一份专业、客观、可直接提交给业务人员的审查报告。

要求：
1. 用中文输出。
2. 必须基于输入数据，不要编造不存在的事实。
3. 严格按以下三个一级标题输出：
   一、基本信息
   二、匹配情况说明
   三、可疑交易分析与建议
4. “基本信息”要覆盖审查文档数、流水笔数、客户数、匹配条数、命中客户数、审查金额等。
5. “匹配情况说明”要解释匹配结果分布、主要命中对象、金额情况，并说明数据结论的依据。
6. “可疑交易分析与建议”必须优先基于整张标准化流水表进行分析，而不是仅基于名单匹配命中的流水；同时可结合匹配结果判断风险关注重点。
7. 行文采用审计报告风格：客观、克制、避免口语化，不使用 Markdown 表格，不使用虚构案例。
8. 每个结论尽量带上数量、金额、时间范围或样本对象；如果输入未提供，就明确说明“根据当前数据无法进一步确认”。
9. 输出为纯文本，适合直接展示在界面中，也适合导出为正式报告文档。
"""


QA_SYSTEM_PROMPT = """你是一个审计问答助手。

你的任务是根据用户问题和给定的最终流水表/匹配详情上下文，提供准确、可追溯的回答。

要求：
1. 用中文输出。
2. 只能根据提供的上下文回答，不要编造数据。
3. 回答要先给结论，再补充依据。
4. 若上下文不足以回答，要明确说明“根据当前最终表格数据，无法确定”，并说明缺了什么。
5. 如引用记录，尽量带上匹配用户、对手名、金额、交易时间、流水行号。
"""


class AuditAgent:
    """基于 OpenAI 兼容接口的报告与问答助手。"""

    def __init__(self, config=None):
        self.config = config or get_config()
        self.api_url = self.config.llm_url.rstrip("/")
        self.model = self.config.llm_model
        self.api_key = self.config.llm_api_key
        self.timeout = self.config.llm_timeout
        self.session = requests.Session()
        # 避免 requests 继承 Windows 系统代理，导致直连 LLM 接口前就在本地代理层失败。
        self.session.trust_env = False

    def is_available(self) -> bool:
        return bool(self.api_key and self.api_url and self.model)

    def generate_report(self, task_title: str, task_id: str, review_result) -> str:
        payload = self._build_report_payload(task_title, task_id, review_result)
        user_message = json.dumps(payload, ensure_ascii=False, indent=2)
        return self._chat(REPORT_SYSTEM_PROMPT, user_message, max_tokens=2200)

    def answer_question(self, question: str, flow_excel_path: str, review_result=None) -> str:
        context_rows = self._select_relevant_rows(question, flow_excel_path)
        payload = {
            "question": question,
            "review_summary": {
                "total_customers": getattr(review_result, "total_customers", 0) if review_result else 0,
                "matched_customers": getattr(review_result, "matched_customers", 0) if review_result else 0,
                "total_matches": getattr(review_result, "total_matches", 0) if review_result else 0,
                "total_amount": getattr(review_result, "total_amount_formatted", "") if review_result else "",
            },
            "retrieved_rows": context_rows,
        }
        user_message = json.dumps(payload, ensure_ascii=False, indent=2)
        return self._chat(QA_SYSTEM_PROMPT, user_message, max_tokens=1500)

    def _chat(self, system_prompt: str, user_message: str, max_tokens: int = 1500) -> str:
        if not self.is_available():
            raise RuntimeError("未配置可用的大模型 API，请先在设置中填写 API 地址、模型和 Key。")

        url = f"{self.api_url}/chat/completions"
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_message},
            ],
            "temperature": 0.2,
            "max_tokens": max_tokens,
        }

        try:
            response = self.session.post(url, headers=headers, json=payload, timeout=self.timeout)
            response.raise_for_status()
            data = response.json()
            message = data["choices"][0]["message"]
            content = (
                message.get("content")
                or message.get("reasoning_content")
                or message.get("reasoning")
            )
            if not content:
                raise RuntimeError(f"大模型返回空内容: {json.dumps(message, ensure_ascii=False)[:500]}")
            return str(content).strip()
        except Exception as exc:
            logger.warning("AuditAgent request failed: %s", exc)
            raise RuntimeError(f"调用大模型失败：{exc}")

    def _select_relevant_rows(self, question: str, flow_excel_path: str) -> List[Dict]:
        rows = self._load_final_rows(flow_excel_path)
        if not rows:
            return []

        keywords = [part.strip() for part in question.replace("，", " ").replace("。", " ").replace("?", " ").replace("？", " ").split() if part.strip()]
        if not keywords:
            return rows[:20]

        scored = []
        for row in rows:
            text = " ".join(str(v) for v in row.values()).lower()
            score = sum(1 for keyword in keywords if keyword.lower() in text)
            if score > 0:
                scored.append((score, row))

        if not scored:
            return rows[:20]

        scored.sort(key=lambda item: item[0], reverse=True)
        return [row for _, row in scored[:20]]

    def _load_final_rows(self, excel_path: str) -> List[Dict]:
        path = Path(excel_path)
        if not path.exists():
            return []

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet_name = "匹配详情" if "匹配详情" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if not headers:
                return []

            header_names = [str(cell).strip() if cell is not None else "" for cell in headers]
            data = []
            for row in rows_iter:
                item = {}
                for idx, header in enumerate(header_names):
                    if header:
                        item[header] = "" if idx >= len(row) or row[idx] is None else str(row[idx]).strip()
                if any(item.values()):
                    data.append(item)
            return data
        finally:
            wb.close()

    def _build_report_payload(self, task_title: str, task_id: str, review_result) -> Dict:
        processed_rows = self._load_processed_rows(getattr(review_result, "flow_excel_path", ""))
        matched_rows = [row for row in processed_rows if self._normalize_text(self._get_field(row, "匹配用户"))]
        analysis_rows = processed_rows or matched_rows

        if not matched_rows:
            matched_rows = [
                {
                    "匹配用户": match.customer_name,
                    "匹配类型": match.match_type,
                    "匹配度": match.confidence,
                    "来源文件": match.source_file,
                    "交易时间": match.transaction_time,
                    "交易对手名": match.counterparty_name,
                    "交易对手账号": match.counterparty_account,
                    "金额": match.amount,
                    "摘要": match.summary,
                    "流水行号": match.row_index,
                }
                for match in getattr(review_result, "matches", [])
            ]
        if not analysis_rows:
            analysis_rows = matched_rows

        suspicious = self._analyze_suspicious_patterns(analysis_rows)
        basic_info = self._build_basic_info(task_title, task_id, review_result, processed_rows, matched_rows)
        match_summary = self._build_match_summary(review_result, matched_rows)

        return {
            "report_goal": "生成员工-客户金额往来审查报告文档正文",
            "analysis_scope": {
                "full_flow_table_rows": len(analysis_rows),
                "matched_rows": len(matched_rows),
                "rule_note": "可疑交易分析基于整张标准化流水表；匹配情况说明基于名单审查命中结果。"
            },
            "writing_rules": [
                "严格使用三个一级标题：一、基本信息；二、匹配情况说明；三、可疑交易分析与建议。",
                "语言保持正式、客观、简洁，适合审计或合规汇报材料。",
                "仅依据输入数据下结论，不得补充输入中不存在的事实。",
                "金额统一保留两位小数并写明单位；数量需明确“份/笔/条/个”等单位。",
                "可疑交易分析默认围绕整张标准化流水表展开，不得错误写成仅基于名单命中流水。",
                "若某类异常未发现，应直接写明“根据当前规则和数据，暂未发现明显异常”。",
            ],
            "basic_info": basic_info,
            "match_summary": match_summary,
            "suspicious_transactions": suspicious,
        }

    def _load_processed_rows(self, excel_path: str) -> List[Dict]:
        path = Path(excel_path)
        if not path.exists():
            return []

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if not headers:
                return []

            header_names = [str(cell).strip() if cell is not None else "" for cell in headers]
            data = []
            for row_index, row in enumerate(rows_iter, start=2):
                item = {"流水行号": row_index}
                for idx, header in enumerate(header_names):
                    if not header:
                        continue
                    value = row[idx] if idx < len(row) else ""
                    item[header] = "" if value is None else str(value).strip()
                if any(str(value).strip() for key, value in item.items() if key != "流水行号"):
                    data.append(item)
            return data
        finally:
            wb.close()

    def _build_basic_info(
        self,
        task_title: str,
        task_id: str,
        review_result,
        processed_rows: List[Dict],
        matched_rows: List[Dict],
    ) -> Dict:
        unique_files = sorted({
            self._normalize_text(self._get_field(row, "来源文件"))
            for row in processed_rows
            if self._normalize_text(self._get_field(row, "来源文件"))
        })

        total_flow_amount = self._sum_amounts(processed_rows)
        matched_amount = self._sum_amounts(matched_rows)
        time_range = self._get_time_range(processed_rows)

        return {
            "task_title": task_title or getattr(review_result, "review_id", ""),
            "task_id": task_id or getattr(review_result, "review_id", ""),
            "review_time": getattr(review_result, "review_time", ""),
            "flow_excel_path": getattr(review_result, "flow_excel_path", ""),
            "reviewed_document_count": len(unique_files),
            "reviewed_documents": unique_files[:20],
            "reviewed_flow_count": len(processed_rows),
            "matched_flow_count": len(matched_rows),
            "customer_count": getattr(review_result, "total_customers", 0),
            "matched_customer_count": getattr(review_result, "matched_customers", 0),
            "matched_record_count": getattr(review_result, "total_matches", len(matched_rows)),
            "total_flow_amount": self._format_amount(total_flow_amount),
            "matched_amount": self._format_amount(matched_amount),
            "review_amount": getattr(review_result, "total_amount_formatted", self._format_amount(total_flow_amount)),
            "time_range": time_range,
            "writeback_status": "失败" if getattr(review_result, "writeback_error", "") else "成功",
            "writeback_error": getattr(review_result, "writeback_error", ""),
        }

    def _build_match_summary(self, review_result, matched_rows: List[Dict]) -> Dict:
        customer_counter: Counter = Counter()
        customer_amount: defaultdict = defaultdict(float)
        match_type_counter: Counter = Counter()
        confidence_values: List[int] = []

        for row in matched_rows:
            customer_name = self._normalize_text(self._get_field(row, "匹配用户"))
            if customer_name:
                customer_counter[customer_name] += 1
                customer_amount[customer_name] += self._parse_amount(self._get_field(row, "金额"))

            match_type = self._normalize_text(self._get_field(row, "匹配类型"))
            if match_type:
                match_type_counter[match_type] += 1

            confidence = self._parse_int(self._get_field(row, "匹配度"))
            if confidence is not None:
                confidence_values.append(confidence)

        top_customers = []
        for customer_name, count in customer_counter.most_common(10):
            top_customers.append({
                "customer_name": customer_name,
                "match_count": count,
                "match_amount": self._format_amount(customer_amount.get(customer_name, 0.0)),
            })

        samples = []
        for row in matched_rows[:20]:
            samples.append({
                "流水行号": row.get("流水行号", ""),
                "匹配用户": self._get_field(row, "匹配用户"),
                "交易时间": self._get_field(row, "交易时间"),
                "交易对手名": self._get_field(row, "交易对手名"),
                "金额": self._get_field(row, "金额"),
                "摘要": self._get_field(row, "摘要"),
            })

        avg_confidence = round(sum(confidence_values) / len(confidence_values), 2) if confidence_values else None

        return {
            "matched_customer_count": getattr(review_result, "matched_customers", 0),
            "matched_record_count": getattr(review_result, "total_matches", len(matched_rows)),
            "matched_amount": self._format_amount(self._sum_amounts(matched_rows)),
            "match_type_distribution": dict(match_type_counter),
            "average_confidence": avg_confidence,
            "top_customers": top_customers,
            "matched_time_range": self._get_time_range(matched_rows),
            "sample_records": samples,
        }

    def _analyze_suspicious_patterns(self, matched_rows: List[Dict]) -> Dict:
        short_interval_groups = self._find_short_interval_groups(matched_rows)
        repeated_counterparties = self._find_repeated_counterparties(matched_rows)
        night_transactions = self._find_night_transactions(matched_rows)
        repeated_same_amount = self._find_same_amount_repeats_within_day(matched_rows)

        return {
            "rules": {
                "short_interval_same_counterparty": "同一交易对手在30分钟内发生2笔及以上交易",
                "repeated_same_counterparty": "同一交易对手累计发生3笔及以上交易",
                "same_amount_within_day": "1天内与同一交易对手发生2笔及以上相同金额交易",
                "night_transactions": "交易时间落在22:00至次日06:00之间",
            },
            "short_interval_same_counterparty": {
                "count": len(short_interval_groups),
                "summary": self._summarize_case_amounts(short_interval_groups),
                "top_cases": short_interval_groups[:8],
            },
            "repeated_same_counterparty": {
                "count": len(repeated_counterparties),
                "summary": self._summarize_case_amounts(repeated_counterparties),
                "top_cases": repeated_counterparties[:8],
            },
            "same_amount_within_day": {
                "count": len(repeated_same_amount),
                "summary": self._summarize_case_amounts(repeated_same_amount),
                "top_cases": repeated_same_amount[:8],
            },
            "night_transactions": {
                "count": len(night_transactions),
                "summary": self._summarize_night_transactions(night_transactions),
                "top_cases": night_transactions[:10],
            },
        }

    def _find_short_interval_groups(self, matched_rows: List[Dict]) -> List[Dict]:
        grouped: defaultdict = defaultdict(list)
        for row in matched_rows:
            tx_time = self._parse_datetime(self._get_field(row, "交易时间"))
            if not tx_time:
                continue
            counterparty = self._normalize_text(self._get_field(row, "交易对手名"))
            if not counterparty:
                continue
            grouped[counterparty].append((tx_time, row))

        cases = []
        for counterparty, items in grouped.items():
            items.sort(key=lambda item: item[0])
            cluster: List[Tuple[datetime, Dict]] = []
            for tx_time, row in items:
                if not cluster:
                    cluster = [(tx_time, row)]
                    continue
                last_time = cluster[-1][0]
                if tx_time - last_time <= timedelta(minutes=30):
                    cluster.append((tx_time, row))
                else:
                    if len(cluster) >= 2:
                        cases.append(self._build_cluster_case(counterparty, cluster))
                    cluster = [(tx_time, row)]
            if len(cluster) >= 2:
                cases.append(self._build_cluster_case(counterparty, cluster))

        cases.sort(key=lambda item: (item["transaction_count"], item["total_amount_value"]), reverse=True)
        for item in cases:
            item.pop("total_amount_value", None)
        return cases

    def _build_cluster_case(self, counterparty: str, cluster: List[Tuple[datetime, Dict]]) -> Dict:
        total_amount = sum(self._parse_amount(self._get_field(row, "金额")) for _, row in cluster)
        return {
            "counterparty_name": counterparty,
            "transaction_count": len(cluster),
            "time_window": f"{cluster[0][0].strftime('%Y-%m-%d %H:%M:%S')} 至 {cluster[-1][0].strftime('%Y-%m-%d %H:%M:%S')}",
            "total_amount": self._format_amount(total_amount),
            "total_amount_value": total_amount,
            "sample_rows": [row.get("流水行号", "") for _, row in cluster[:10]],
            "sample_summaries": [self._get_field(row, "摘要") for _, row in cluster[:3] if self._get_field(row, "摘要")],
        }

    def _find_repeated_counterparties(self, matched_rows: List[Dict]) -> List[Dict]:
        grouped: defaultdict = defaultdict(list)
        for row in matched_rows:
            counterparty = self._normalize_text(self._get_field(row, "交易对手名"))
            if not counterparty:
                continue
            grouped[counterparty].append(row)

        cases = []
        for counterparty, rows in grouped.items():
            if len(rows) < 3:
                continue
            total_amount = self._sum_amounts(rows)
            cases.append({
                "counterparty_name": counterparty,
                "transaction_count": len(rows),
                "total_amount": self._format_amount(total_amount),
                "date_range": self._get_time_range(rows),
                "sample_rows": [row.get("流水行号", "") for row in rows[:10]],
            })

        cases.sort(key=lambda item: (item["transaction_count"], self._amount_to_value(item["total_amount"])), reverse=True)
        return cases

    def _find_same_amount_repeats_within_day(self, matched_rows: List[Dict]) -> List[Dict]:
        grouped: defaultdict = defaultdict(list)
        for row in matched_rows:
            tx_time = self._parse_datetime(self._get_field(row, "交易时间"))
            counterparty = self._normalize_text(self._get_field(row, "交易对手名"))
            amount_value = self._parse_amount(self._get_field(row, "金额"))
            if not tx_time or not counterparty or amount_value <= 0:
                continue
            grouped[(counterparty, tx_time.date().isoformat(), round(amount_value, 2))].append((tx_time, row))

        cases = []
        for (counterparty, tx_date, amount_value), items in grouped.items():
            if len(items) < 2:
                continue
            items.sort(key=lambda item: item[0])
            cases.append({
                "counterparty_name": counterparty,
                "transaction_date": tx_date,
                "same_amount": self._format_amount(amount_value),
                "transaction_count": len(items),
                "time_points": [item[0].strftime("%H:%M:%S") for item in items[:8]],
                "sample_rows": [row.get("流水行号", "") for _, row in items[:10]],
                "sample_summaries": [self._get_field(row, "摘要") for _, row in items[:3] if self._get_field(row, "摘要")],
                "total_amount": self._format_amount(amount_value * len(items)),
            })

        cases.sort(key=lambda item: (item["transaction_count"], self._amount_to_value(item["total_amount"])), reverse=True)
        return cases

    def _find_night_transactions(self, matched_rows: List[Dict]) -> List[Dict]:
        rows = []
        for row in matched_rows:
            tx_time = self._parse_datetime(self._get_field(row, "交易时间"))
            if not tx_time:
                continue
            if tx_time.hour >= 22 or tx_time.hour < 6:
                rows.append({
                    "流水行号": row.get("流水行号", ""),
                    "transaction_time": tx_time.strftime("%Y-%m-%d %H:%M:%S"),
                    "counterparty_name": self._get_field(row, "交易对手名"),
                    "amount": self._get_field(row, "金额"),
                    "summary": self._get_field(row, "摘要"),
                })
        return rows

    def _summarize_case_amounts(self, cases: List[Dict]) -> Dict:
        total_amount = 0.0
        total_transactions = 0
        counterparties = set()
        for case in cases:
            total_amount += self._amount_to_value(case.get("total_amount", ""))
            total_transactions += int(case.get("transaction_count", 0) or 0)
            name = self._normalize_text(case.get("counterparty_name", ""))
            if name:
                counterparties.add(name)
        return {
            "total_amount": self._format_amount(total_amount),
            "total_transactions": total_transactions,
            "counterparty_count": len(counterparties),
        }

    def _summarize_night_transactions(self, rows: List[Dict]) -> Dict:
        counterparty_counter: Counter = Counter()
        total_amount = 0.0
        for row in rows:
            name = self._normalize_text(row.get("counterparty_name", ""))
            if name:
                counterparty_counter[name] += 1
            total_amount += self._parse_amount(row.get("amount", ""))
        return {
            "total_amount": self._format_amount(total_amount),
            "counterparty_count": len(counterparty_counter),
            "top_counterparties": [
                {"counterparty_name": name, "transaction_count": count}
                for name, count in counterparty_counter.most_common(5)
            ],
        }

    @staticmethod
    def _get_field(row: Dict, field: str) -> str:
        aliases = {
            "交易对手名": ["交易对手名", "对手名"],
            "交易对手账号": ["交易对手账号", "对手账号"],
            "匹配用户": ["匹配用户"],
            "匹配类型": ["匹配类型"],
            "匹配度": ["匹配度"],
            "来源文件": ["来源文件"],
            "交易时间": ["交易时间"],
            "金额": ["金额"],
            "摘要": ["摘要"],
        }
        for key in aliases.get(field, [field]):
            value = row.get(key, "")
            if value not in (None, ""):
                return str(value).strip()
        return ""

    @staticmethod
    def _normalize_text(value: object) -> str:
        return str(value or "").strip()

    @staticmethod
    def _parse_amount(amount_value: object) -> float:
        if amount_value is None:
            return 0.0
        amount_str = str(amount_value).strip()
        if not amount_str:
            return 0.0
        clean = (
            amount_str.replace(",", "")
            .replace("￥", "")
            .replace("¥", "")
            .replace("元", "")
            .replace("+", "")
        )
        try:
            return abs(float(clean))
        except (ValueError, TypeError):
            return 0.0

    def _sum_amounts(self, rows: List[Dict]) -> float:
        return sum(self._parse_amount(self._get_field(row, "金额")) for row in rows)

    @staticmethod
    def _format_amount(value: float) -> str:
        return f"¥{value:,.2f}"

    def _get_time_range(self, rows: List[Dict]) -> Dict[str, str]:
        times = [
            self._parse_datetime(self._get_field(row, "交易时间"))
            for row in rows
        ]
        valid_times = [item for item in times if item]
        if not valid_times:
            return {"start": "", "end": ""}
        return {
            "start": min(valid_times).strftime("%Y-%m-%d %H:%M:%S"),
            "end": max(valid_times).strftime("%Y-%m-%d %H:%M:%S"),
        }

    @staticmethod
    def _parse_datetime(value: object) -> Optional[datetime]:
        text = str(value or "").strip()
        if not text:
            return None

        candidates = [
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M",
            "%Y-%m-%d",
            "%Y/%m/%d",
        ]
        for fmt in candidates:
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            return None

    @staticmethod
    def _parse_int(value: object) -> Optional[int]:
        text = str(value or "").strip()
        if not text:
            return None
        try:
            return int(float(text))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _amount_to_value(text: str) -> float:
        clean = str(text or "").replace("¥", "").replace(",", "").strip()
        try:
            return float(clean)
        except (ValueError, TypeError):
            return 0.0
