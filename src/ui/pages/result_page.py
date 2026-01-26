# -*- coding: utf-8 -*-
"""
Result page - review results display and export
"""
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
    QPushButton, QMessageBox, QFileDialog
)

from ..widgets import Card, ResultTable
from ..styles import COLORS
from ...config import get_config


class ResultPage(QWidget):
    """
    Result page for displaying review results
    - Match details table
    - Export functionality
    """
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.config = get_config()
        self.result = None
        self._setup_ui()
    
    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(24)
        
        # Page title
        header_layout = QHBoxLayout()
        
        title_layout = QVBoxLayout()
        title = QLabel("审查结果")
        title.setObjectName("title")
        title.setStyleSheet(f"""
            font-size: 24px;
            font-weight: bold;
            color: {COLORS['text_primary']};
        """)
        title_layout.addWidget(title)
        
        self.subtitle = QLabel("审查完成，查看匹配结果")
        self.subtitle.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 14px;")
        title_layout.addWidget(self.subtitle)
        
        header_layout.addLayout(title_layout)
        header_layout.addStretch()
        
        layout.addLayout(header_layout)
        
        # Result table
        table_card = Card()
        table_card.add_title("匹配明细")
        
        self.result_table = ResultTable()
        table_card.layout.addWidget(self.result_table)
        
        layout.addWidget(table_card, 1)  # Stretch to fill
        
        # Bottom action bar
        action_layout = QHBoxLayout()
        
        action_layout.addStretch()

        self.export_btn = QPushButton("导出Excel")
        self.export_btn.setObjectName("secondary_btn")
        self.export_btn.clicked.connect(self._export_excel)
        action_layout.addWidget(self.export_btn)
        
        self.new_audit_btn = QPushButton("新建审查")
        self.new_audit_btn.setObjectName("secondary_btn")
        action_layout.addWidget(self.new_audit_btn)
        
        layout.addLayout(action_layout)
    
    def set_review_result(self, result) -> None:
        """
        设置审查结果（来自 Reviewer）
        
        Args:
            result: ReviewResult 对象
        """
        self.result = result
        
        # 转换匹配记录格式并更新表格
        matches_data = [m.to_dict() for m in result.matches]
        self.result_table.set_data(matches_data)
        
        # 更新副标题
        if result.total_matches == 0:
            self.subtitle.setText("未发现匹配记录")
        else:
            self.subtitle.setText(f"发现 {result.total_matches} 条匹配记录")
    
    def _export_excel(self) -> None:
        """Export review results to Excel"""
        if not self.result:
            QMessageBox.warning(self, "无数据", "没有可导出的审查结果")
            return
        
        default_name = f"审查结果_{self.result.review_id}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出审查结果",
            str(self.config.reports_folder / default_name),
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
            from pathlib import Path as SysPath
            
            wb = Workbook()
            ws = wb.active
            ws.title = "匹配明细"
            
            headers = ["匹配用户", "来源文件", "交易时间", "对手名", "对手账号", "金额", "摘要"]
            widths = [15, 30, 20, 18, 20, 12, 30]
            
            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]
            
            for row_idx, match in enumerate(self.result.matches, 2):
                data = [
                    match.customer_name,
                    SysPath(match.source_file).name,
                    match.transaction_time,
                    match.counterparty_name,
                    match.counterparty_account,
                    match.amount,
                    match.summary,
                ]
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
            
            ws.freeze_panes = 'A2'
            wb.save(file_path)
            
            QMessageBox.information(
                self, "导出成功",
                f"审查结果已导出到:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.warning(
                self, "导出失败",
                f"导出审查结果时出错:\n{str(e)}"
            )
    
    def clear(self) -> None:
        """Clear results"""
        self.result = None
        self.result_table.clear()
